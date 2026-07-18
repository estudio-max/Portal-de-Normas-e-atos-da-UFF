# -*- coding: utf-8 -*-
"""
Extrator de Atos do Boletim de Serviço da UFF
==============================================

Lê os PDFs do Boletim de Serviço (https://boletimdeservico.uff.br) e produz
uma base estruturada (CSV/XLSX/JSON) com um registro por ato administrativo,
incluindo:

  - tipo, órgão/sigla, número, data do ato e ementa
  - número do Boletim, data do Boletim, seção e página
  - processos SEI citados (23069.XXXXXX/AAAA-DD)
  - código SEI do documento (código verificador)
  - relações entre atos: ALTERA / REVOGA / SUBSTITUI / RETIFICA /
    REPUBLICA / TORNA SEM EFEITO / PRORROGA / CITA

Uso:
    python extrair_boletim.py                 # processa todos os PDFs da pasta ./boletins
    python extrair_boletim.py 53-26.pdf 54-26.pdf 55-26.pdf
    python extrair_boletim.py --pasta "C:/.../Downloads"

Requisitos: PyMuPDF (pip install pymupdf) e openpyxl (pip install openpyxl).
"""

import sys
import os
import re
import json
import glob
import argparse
import unicodedata
from datetime import datetime, date

import fitz  # PyMuPDF

# --------------------------------------------------------------------------- #
# Constantes / vocabulário
# --------------------------------------------------------------------------- #

MESES = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8, "setembro": 9,
    "outubro": 10, "novembro": 11, "dezembro": 12,
}

# Tipos de ato reconhecidos no início de um título.
TIPOS = [
    "DETERMINAÇÃO DE SERVIÇO",
    "INSTRUÇÃO NORMATIVA",
    "NORMA DE SERVIÇO",
    "ORDEM DE SERVIÇO",
    "RESOLUÇÃO",
    "PORTARIA",
    "DECISÃO",
    "DELIBERAÇÃO",
    "EDITAL",
    "COMUNICADO",
    "RESUMO DE DESPACHOS E DECISÕES",
    "RESUMO DE DESPACHOS",
]
# Regex alternativa de tipos (mais longos primeiro p/ casar o mais específico).
# Tolera o typo real da UFF "DETERMINAÇÃO DE SERIIÇO" (190-20.pdf) — 1 a 3
# letras quaisquer entre "SER" e "ÇO"; canonizado depois em canon_tipo().
TIPOS_RE = "|".join(re.escape(t) for t in sorted(TIPOS, key=len, reverse=True))
TIPOS_RE += r"|DETERMINA[ÇC][ÃA]O DE SER[A-ZÀ-Ú]{1,3}[ÇC]O"


def canon_tipo(t):
    """Normaliza o tipo casado: caixa alta e correção de typos conhecidos."""
    t = t.upper()
    if t.startswith("DETERMINA") and t != "DETERMINAÇÃO DE SERVIÇO":
        return "DETERMINAÇÃO DE SERVIÇO"
    return t

# Abreviação de "número" no texto do BS: "Nº", "N°", "No", "N.º" ou "Nº.". A
# ORDEM varia entre boletins/anos — muitos usam PONTO ANTES do símbolo º
# ("PORTARIA N.º 54.919"), não depois. Um regex "[ºo°]?\.?" só aceita a ordem
# símbolo-depois-ponto e por isso IGNORAVA por completo títulos assim, jogando
# o ato inteiro (e suas pessoas/SIAPEs) dentro do trecho do ato vizinho que
# por acaso casou (bug real: Portaria 55.002/2015 "herdou" 71 SIAPEs de outras
# portarias do mesmo boletim). Aceita qualquer ordem/repetição de ./º/°/o, e
# também "O" maiúsculo ("TEC/TCE NO 015" — "Nº" digitado/OCRizado como "NO").
_ORD = r"[.ºo°O]{0,3}"

# Cauda comum do título: número + data. Tolera as variações reais do BS:
#   "Nº 052, DE 13 DE OUTUBRO DE 2020"        (padrão)
#   "Nº 03/2020 - NITERÓI, 09 DE OUTUBRO DE 2020"  (num/ano + cidade, sem "DE" antes do dia)
#   "N° 14 DE 13 DE OUTUBRO 2020"             (sem "DE" antes do ano — typo comum)
_TIT_NUM_DATA = (
    # sufixo do número ("004 AR") não pode engolir o "DE" da data. O "N" do "Nº"
    # aceita minúsculo mesmo com TIPO em caixa alta ("SERVIÇO — HUAP - nº 58").
    r"[Nn]%s\s*(?P<numero>[\d\.]+(?:\s*(?!DE\b)[A-Z]{1,4})?)(?:\s*/\s*\d{2,4})?\s*,?\s*"
    r"(?:[-–]\s*[A-ZÀ-Ú][A-Za-zÀ-ÿ ]{1,30},?\s*)?"
    # Conectores de data podem vir minúsculos: as portarias da Reitoria (SIGA-EX)
    # saem como "Nº 68.884 de 4 de fevereiro de 2026". O TIPO continua exigido em
    # MAIÚSCULO (ou ancorado no marcador SIGA, abaixo), então citações minúsculas
    # no corpo ("portaria nº 8858, de ...") NÃO viram falso título.
    # "de" antes do mês também pode faltar ("de 10 março de 2022").
    r"(?:[Dd][Ee]\s+)?(?P<dia>\d{1,2})\s+(?:[Dd][Ee]\s+)?(?P<mes>[A-Za-zçÇãÃéíóúâêôõ]+)\s+(?:[Dd][Ee]\s+)?(?P<ano>\d{4})"
) % _ORD

# Título de um ato. Ex.:
#   DETERMINAÇÃO DE SERVIÇO COLUNI/UFF Nº. 20, DE 12 DE JUNHO DE 2026
#   PORTARIA Nº 1004, DE 10 DE JUNHO DE 2026
#   RESOLUÇÃO CEPEX/UFF Nº 004 AR, DE 10 DE JUNHO DE 2026
#   PORTARIA N.º 54.919 de 11 de novembro de 2015
#   DETERMINAÇÃO DE SERVIÇO DDRH, Nº. 068 de 20 de julho de 2010  (vírgula antes do Nº)
#   DETERMINAÇÃO DE SERVIÇO — HUAP - nº 58, de 26 de setembro de 2000  (travessão)
TITULO_RE = re.compile(
    r"(?P<tipo>%s)\s+"
    r"(?P<orgao>[A-ZÀ-Ú0-9/().\-–— ]{0,40}?)?,?\s*"
    % TIPOS_RE + _TIT_NUM_DATA
)

# Portarias emitidas pelo SIGA em ALGUNS anos (ex.: 2020) saem em Title Case
# ("Portaria Nº 67.634 de 16 de outubro de 2020") — invisíveis ao TITULO_RE,
# que exige TIPO em CAIXA ALTA. A âncora segura é o marcador de documento SIGA
# ("UFFPOR202067634A") na linha imediatamente anterior — citações no corpo do
# texto nunca têm esse marcador, então não há falso positivo.
TITULO_SIGA_RE = re.compile(
    r"UFF[A-Z]{3}\d{6,}[A-Z]?\s*\n\s*"
    r"(?P<tipo>Portaria|PORTARIA)(?P<orgao>)\s+" + _TIT_NUM_DATA
)

# Marcador de fim/início de ato usado em todo o corpus: OU o separador de
# hashes entre atos na mesma página, OU o cabeçalho/rodapé de página
# ("PÁG. 09") quando o ato começa bem no topo da página seguinte. Ancora com
# segurança formatos CURTOS de título que não têm data por extenso na mesma
# linha — comuns em Decisões/Resoluções de colegiados (CEP/CUV) em anos mais
# antigos: "DECISÃO N.º 026/2012", "RESOLUÇÃO 18/2002" (às vezes sem "Nº"
# nenhum). Como citações no corpo NUNCA vêm logo após um desses marcadores nem
# são seguidas de linha em branco, não há risco de falso título.
#
# O separador tem CINCO hashes em 2002-2003 e SEIS de 2004 em diante — medido
# no corpus: 2002 tem 90 sequências de 5 contra 14 de 6; 2003, 208 de 5 e zero
# de 6; 2004+, só 6. Exigir seis (como era até 16/07/2026) cegava o extrator
# justamente nos anos de cinco. Aceitar 5-ou-6, medido sobre os anos INTEIROS:
# 2002 +189 atos (+19%), 2003 +121 (+15%), 2004 +9, e ZERO atos perdidos em
# qualquer ano de 2001 a 2026. Os recuperados são Decisões reais do CEP em
# numeração sequencial (642, 643, 645...), que é o formato curto que este
# regex existe para ancorar.
# 2001 ganha ZERO aqui, e não é falha do regex: naquele ano o BS é digitalizado
# e o OCR de época transformou todo "# # # # #" em lixo ("HNHUA", "hehe") — não
# há hash nenhum no texto para casar. Só um re-OCR recupera aquele ano.
# Não aceite QUATRO: as 21 sequências de 4 medidas em 2002 não foram
# investigadas e o ganho não compensa o risco de casar coisa que não é
# separador.
_HASH_SEP = r"#(?:\s*#){4,5}"
_ATO_BOUNDARY = r"(?:%s|P[ÁA]G\.?\s*0?\d+)" % _HASH_SEP
TITULO_CURTO_RE = re.compile(
    r"%s\s*\n\s*"
    r"(?P<tipo>DECISÃO|RESOLUÇÃO)(?P<orgao>)\s+"
    r"(?:[Nn]%s\s*)?(?P<numero>[\d\.]+)\s*/\s*(?P<ano>\d{2,4})\s*\.?"
    r"(?P<dia>)(?P<mes>)"
    r"(?=[ \t]*\n[ \t]*\n)"
    % (_ATO_BOUNDARY, _ORD)
)

# Cabeçalhos de documentos publicados no BS que NÃO são "atos" no vocabulário
# do app (extrato de contrato/convênio, termo de homologação/adesão, ata de
# registro de preços — não têm tipo_ato próprio) mas SÃO um documento novo
# começando. Sem reconhecê-los, o corpo do ATO ANTERIOR "engolia" o documento
# inteiro até o próximo título REAL (achado 12/07/2026: DTS PROAD 61/2025
# ganhou um "prazo de 12 meses" que pertencia a um Extrato de Instrumento
# Convenial sobre um terminal portuário, publicado logo depois no boletim —
# 557 atos afetados no corpus, 91% só de "Extrato de Instrumento Convenial").
# Só serve de FRONTEIRA (corta o corpo do ato anterior aqui); não vira ato
# próprio — ver uso de BOUNDARY_NAO_ATO_RE em parse_pdf().
#
# 15/07/2026: mesma classe de bug, achada por outro caminho — seções sem
# cabeçalho de ato próprio ("Resumo de Despachos e Decisões", "Alteração de
# Carga Horária", "Auxílio Funeral") ficavam penduradas no ato anterior. Caso-
# prova: a Portaria 64.814/2019 (Comissão Interna de Conservação de Energia)
# nomeia 9 servidores; o extrator lhe atribuía 22 — os 13 extras eram só
# alterações de carga horária publicadas na sequência do mesmo boletim.
# Medido nos 78.994 atos do corpus: 2.414 atos (3,1%) engolem uma dessas
# seções. "RESUMO DE DESPACHOS E DECISÕES" sozinho cobre 2.343 (97,1%).
#
# SÓ este marcador entrou. "AUXÍLIO FUNERAL" e "ALTERAÇÃO DE CARGA HORÁRIA"
# foram cogitados e DESCARTADOS — ver docs/GUIA-EXTRACAO-BS.md:
#   - "Auxílio Funeral" às vezes imprime como "Assunto: Auxílio Funeral."
#     (Title Case, não caixa alta), então só pega com match
#     case-insensitive — e caiu junto lixo: "com alteração de carga
#     horária" aparece dentro do dispositivo de atos REAIS (DTS 16/2024,
#     sobre mudança de regime de trabalho), então case-insensitive teria
#     decapitado atos legítimos por 2 casos a mais só.
#   - A prova que fechou a decisão: um ato real de 2015 cita "Autorizo o
#     cancelamento dos efeitos do Resumo de Despachos e Decisões n°
#     62/2012" em Title Case, como referência — não como cabeçalho. Se o
#     match fosse case-insensitive, esse ato teria sido cortado ali.
#
# Por isso o match é SEMPRE case-sensitive, igual às fronteiras acima:
# "RESUMO DE DESPACHOS E DECISÕES" imprime em caixa alta pura como título de
# seção (confirmado em amostra de 9 anos, 2005-2026, 248 ocorrências, 1 única
# exceção — a citação Title Case acima, que o case-sensitive já rejeita
# corretamente). Citação em prosa normal não sai em caixa alta.
BOUNDARY_NAO_ATO_RE = re.compile(
    r"(?P<tipo>EXTRATO DE INSTRUMENTO CONVENIAL|EXTRATO DE CONTRATO|EXTRATO DE TERMO ADITIVO"
    r"|EXTRATO DE CONV[ÊE]NIO|TERMO DE HOMOLOGA[ÇC][ÃA]O|TERMO DE ADES[ÃA]O"
    r"|ATA DE REGISTRO DE PRE[ÇC]OS"
    r"|RESUMO DE DESPACHOS E DECIS[ÕO]ES)"
)

# Processo SEI: 23069.166342/2026-40  (aceita espaços no lugar de . / -)
PROC_RE = re.compile(r"23069[.\s]\d{6}[/\s]\d{4}[-\s]\d{2}")
# Código verificador SEI: "SEI nº 3441183"  ou  "(3442574)"
SEI_DOC_RE = re.compile(r"SEI\s*n%s\s*(\d{6,8})" % _ORD)
SEI_DOC_PAREN_RE = re.compile(r"\((\d{6,8})\)")

# Matrícula SIAPE: "SIAPE 1642620", "Siape nº 1642620", "Matrícula SIAPE nº 2364493".
# Consome também a abreviação de "Matrícula" colada/pontuada ("Mat. SIAPE 123",
# "MATSIAPE 123") — senão o "MAT" era absorvido como sobrenome ("...SANTOS MAT").
SIAPE_RE = re.compile(
    r"(?:\b(?:matr[íi]cula|mat)\.?\s*)?(?:SIAPE|Siape)[:\s]*n?%s\s*(\d{6,7})" % _ORD, re.I)

# Nome da pessoa citada, ancorado na matrícula SIAPE que aparece logo depois.
# _CONNECT: conectores que ficam minúsculos no nome ("de", "da", ...).
# _BLOCK_NOME: verbos/cargos/descritores que NÃO fazem parte do nome (sem acento).
_CONNECT = {"de", "da", "do", "das", "dos", "e"}
_BLOCK_NOME = set((
    "professor professora reitor reitora diretor diretora coordenador coordenadora coord "
    "chefe presidente secretario secretaria tecnico analista assistente magisterio superior "
    "matricula codigo cargo ocupante servidor servidora substituto substituta membro membros "
    "comissao subcomissao universidade federal fluminense ministerio educacao departamento "
    "divisao coordenacao reitoria central documento paragrafo considerando quadro permanente "
    "senhor senhora gabinete vice decano superintendente lotar designar dispensar exonerar "
    "nomear conceder autorizar instituir revogar alterar prorrogar tornar retificar republicar "
    "considerar resolver delegar aprovar homologar redistribuido "
    "professores professoras docente docentes diretores coordenadores coordenadoras chefes presidentes "
    "secretarios tecnicos analistas assistentes servidores substitutos assuntos educacionais "
    "educacional gerais administrativos administrativa academicos academica financeiros "
    "institucional institucionais setor nucleo gerencia assessoria pessoal ensino pesquisa "
    "extensao graduacao licenca afastamento concessao capacitacao "
    # cargos/ocupações — evita capturar o cargo no lugar do nome ("Auxiliar de
    # Enfermagem", "Produtor Cultural"). Só termos que NÃO servem de sobrenome.
    "auxiliar auxiliares enfermeiro enfermeira enfermagem medico medica odontologo odontologa "
    "dentista cirurgiao psicologo psicologa nutricionista farmaceutico farmaceutica bioquimico "
    "fisioterapeuta fonoaudiologo biologo biomedico veterinario zootecnista engenheiro engenheira "
    "arquiteto arquiteta quimico geologo estatistico economista contador contadora administrador "
    "administradora administrativo bibliotecario bibliotecaria documentalista arquivista museologo "
    "jornalista publicitario tradutor tradutora revisor revisora programador programadora operador "
    "operadora motorista telefonista recepcionista almoxarife vigilante porteiro copeiro cozinheiro "
    "cozinheira servente jardineiro eletricista mecanico soldador torneiro marceneiro desenhista "
    "fotografo diagramador instrumentador produtor produtora cultural social sociologo historiador "
    "pedagogo pedagoga artifice contabilidade laboratorio laboratorista "
    # áreas/departamentos e especialidades — em tabelas de colegiado o nome do
    # departamento antecede o nome da pessoa ("Psicologia João Batista..."). Só
    # termos que não servem de sobrenome.
    "sociologia antropologia geografia historia filosofia psicologia pedagogia letras linguistica "
    "comunicacao jornalismo economia administracao direito arquitetura engenharia matematica fisica "
    "quimica biologia geologia computacao informatica ciencia ciencias politica politicas sociais "
    "humanas naturais exatas biologicas medicina odontologia farmacia nutricao fisioterapia "
    "fonoaudiologia veterinaria cirurgia clinica pediatria ginecologia obstetricia cardiologia "
    "neurologia ortopedia radiologia anestesiologia dermatologia oftalmologia urologia patologia "
    "psiquiatria geriatria traumatologia bucomaxilofacial bucomaxilofaciais buco maxilo facial geral "
    # ruído de cabeçalho de tabela ("Nome:") e abreviação de "Matrícula" que colam no nome
    "nome mat "
    # referências a outros atos coladas no nome ("Revogar a DTS GGE/UFF Nº 1..."
    # deixava "Dts Gge" como nome)
    "dts portaria determinacao resolucao edital comunicado instrucao normativa "
    # cabeçalho de quebra de página no meio do ato ("...BOLETIM DE SERVIÇO
    # ANO LX... SEÇÃO IV PÁG." virava "Serviço Ano Lx", "Seção Iv")
    "boletim servico ano secao pag "
    # rótulos de tabela/edital eleitoral e áreas que colam no nome
    "chapa titular titulares suplente suplentes eleitoral local prox analitica "
    "instituto faculdade escola desenvolvimento regional monitoria genetica "
    "ac aa "           # tipo de bolsa (Ampla Concorrência/Ação Afirmativa) em tabelas
    "superintenente"   # typo recorrente de "superintendente" nas DTS da SAEP
).split())
# Classes de letras Latin-1 COMPLETAS: a faixa antiga "à-ú" não cobria ü/û/ý/ÿ
# nem Ü/Þ, o que truncava sobrenomes germânicos ("Frühauf"->"Fr", "SCHMÜTZ"->
# "Schm"). O apóstrofo ('/’) integra sobrenomes ("Sant'Anna", "Dal'Magro",
# "D'Almeida"). A cauda aceita CAIXA MISTA p/ tolerar typo de caps-lock do PDF
# ("LIma"->"Lima", "LIzarbe"->"Lizarbe").
_UP = "A-ZÀ-ÖØ-Þ"          # maiúsculas Latin-1 (pula × U+00D7)
_LO = "a-zà-öø-ÿ"          # minúsculas Latin-1 (pula ÷ U+00F7)
_LET = _UP + _LO
_AP = "'’"
_PALAVRA_NOME = r"[%s][%s]*(?:[%s][%s]+)*" % (_UP, _LET, _AP, _LET)
# NOME_CAPS_RE prioriza CAIXA ALTA (distingue nome de cargo, que vem em title-
# case): exige 2+ maiúsculas iniciais, mas tolera cauda minúscula de typo/OCR
# ("WASSERMAn" -> "Wasserman").
_PALAVRA_CAPS = r"[%s]{2,}[%s]*(?:[%s][%s]+)*" % (_UP, _LET, _AP, _LET)
# Conector: inclui "d'Aquino"/"d'Ávila" inteiros (senão o "d'" corta o sobrenome).
# "d[aeo]s?" e não "de|da|do|das|dos": na alternância, "do" casa antes e
# esconde "dos" (o motor não volta atrás), partindo "Lenin dos Santos Pires"
# em "Lenin do" + "Santos Pires".
_CONN = r"(?:d[aeo]s?|e|d[%s][%s]*)" % (_AP, _LET)
NOME_RE = re.compile(r"%s(?:\s+(?:%s|%s)){1,6}" % (_PALAVRA_NOME, _CONN, _PALAVRA_NOME))
# Run de CAIXA ALTA não aceita conector minúsculo: nome oficial em caps usa
# "DOS/DE" também em caps (casam como palavra), e o minúsculo no meio denuncia
# sigla/rótulo ("Comissão CG/PROX do PPGQ"), que virava nome.
NOME_CAPS_RE = re.compile(r"%s(?:\s+%s){1,6}" % (_PALAVRA_CAPS, _PALAVRA_CAPS))

# Linha de cabeçalho repetida em cada página do ato
HEADER_BS_RE = re.compile(
    r"UNIVERSIDADE FEDERAL FLUMINENSE.{0,5}BOLETIM DE SERVIÇO", re.I)
ANO_NUM_RE = re.compile(r"ANO\s+([IVXLCDM]+)\s*.{0,4}\s*N%s\s*(\d+)" % _ORD, re.I)
DATA_BS_RE = re.compile(r"\b(\d{2}/\d{2}/\d{4})\b")
SECAO_RE = re.compile(r"SEÇÃO\s+([IVX]+)\s+(?:PÁG|P)\.?\s*0?(\d+)", re.I)

# Verbos que indicam relação entre atos (ordem = prioridade de classificação).
# Cada par (rótulo, regex). A busca é ANCORADA no ato citado: para cada
# referência a outro ato, olha-se o texto IMEDIATAMENTE ANTERIOR procurando
# um destes verbos.
VERBOS_RELACAO = [
    ("REVOGA",           re.compile(r"revog", re.I)),
    ("TORNA SEM EFEITO", re.compile(r"torna(r|m)?\s+sem\s+efeito", re.I)),
    ("SUBSTITUI",        re.compile(r"substitu", re.I)),
    ("RETIFICA",         re.compile(r"retific", re.I)),
    ("REPUBLICA",        re.compile(r"republic", re.I)),
    ("ANULA",            re.compile(r"\banul", re.I)),
    ("PRORROGA",         re.compile(r"prorrog", re.I)),
    ("ALTERA",           re.compile(r"alter(a|ar|ação|açao)", re.I)),
    ("CITA",             re.compile(r"conforme|nos\s+termos|com\s+base|com\s+fundamento|"
                                    r"previst|de\s+que\s+trata|estabelecid|delegad", re.I)),
]
# Verbo no início da ementa -> natureza do ato (para classificação rápida)
ACAO_EMENTA_RE = re.compile(
    r"^\s*(Altera|Revoga|Substitui|Retifica|Republica|Designa|Designar|Dispõe|"
    r"Prorroga|Torna|Aprova|Cria|Institui|Estabelece|Nomeia|Exonera|Dispensa|"
    r"Concede|Autoriza|Delega|Constitui|Homologa|Cancela|Suspende|Anula)", re.I)

# Referência a um ato citado dentro do corpo, ex.:
#   "Portaria nº 65.784, de 29/11/2019"
#   "DTS GES/INF/UFF nº 16, de 22 de agosto de 2025"
#   "Resolução CUV nº 026 de 18 de abril de 2017"
REF_TIPOS = (r"Portaria|Resolução|Determinação de Serviço|DTS|Norma de Serviço|"
             r"Instrução Normativa|Decisão|Edital|Ordem de Serviço|Deliberação")
# Aceita as DUAS grafias do número do ato citado:
#   • com marcador "nº":  "DTS GES/INF/UFF nº 16, de 22/08/2025"  (ano opcional /AAAA)
#   • forma COMPACTA sem "nº":  "DTS GHT 07/2023"  (sigla + NN/AAAA) — muito usada
#     em revogações ("Revoga a DTS GHT 07/2023"). Exige "/AAAA" p/ NÃO capturar
#     número solto (artigo, processo, "Lei 8.112/90"); o tipo já vem da whitelist
#     REF_TIPOS e a relação só é registrada se houver verbo (revoga/altera/...).
# Órgão: cada token COMEÇA com letra — senão, na forma compacta "GHT 07/2023" o
# órgão (que aceita dígitos p/ siglas tipo "GES/INF/UFF") engoliria o "0" do
# número ("GHT 0" + "7/2023"). Exigir letra inicial deixa o número intacto.
REF_RE = re.compile(
    r"(?P<tipo>(?i:%s))\s+"
    r"(?P<orgao>[A-ZÀ-Ú][A-ZÀ-Ú0-9/().]{0,24}(?:\s[A-ZÀ-Ú][A-ZÀ-Ú0-9/().]{0,14}){0,3})?\s*"
    r"(?:"
    r"[nN]%s\s*(?P<numero>\d[\d\.]*(?:\s*[A-Z]{1,4})?)(?:\s*/\s*(?P<ano>\d{4}))?"
    r"|"
    r"(?P<numero2>\d{1,4})\s*/\s*(?P<ano2>\d{4})"
    r")" % (REF_TIPOS, _ORD))

# Referência a outro Boletim: "publicada no BS nº 102, de 01/09/2025"
BS_REF_RE = re.compile(r"BS\s*n%s\s*(\d+)\s*,?\s*de\s*(\d{2}/\d{2}/\d{4})" % _ORD, re.I)


# --------------------------------------------------------------------------- #
# Funções utilitárias
# --------------------------------------------------------------------------- #

# Caracteres de controle ilegais (rejeitados por XLSX e indesejados no resto)
CTRL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def limpar(txt):
    """Normaliza espaços/quebras de linha e remove caracteres de controle."""
    txt = CTRL_RE.sub("", txt)
    txt = re.sub(r"[ \t]+", " ", txt)
    txt = re.sub(r"\s*\n\s*", " ", txt)
    return txt.strip()


def normaliza_proc(s):
    """Normaliza um processo SEI para 23069.XXXXXX/AAAA-DD."""
    d = re.sub(r"\D", "", s)
    if len(d) >= 17:
        return f"{d[0:5]}.{d[5:11]}/{d[11:15]}-{d[15:17]}"
    return s.strip()


def _data_valida(iso):
    try:
        datetime.strptime(iso, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def corrige_ano_futuro(data_ato, ano_ato, bs_data):
    """Typo de ano na fonte ("DE 03 DE DEZEMBRO DE 2026" num BS de 2025;
    "Nº 047/2209") deixa o ato datado do futuro — impossível, o boletim só
    publica ato já assinado. A âncora é a data do PRÓPRIO boletim, que vale
    também nos anos antigos do backfill (a data corrente não denunciaria um
    "2022" digitado num boletim de 2021).
    - data_ato depois do BS: tenta o ano do BS e o anterior (ato de dezembro
      publicado em janeiro); fica com o primeiro que não seja futuro. Sem
      candidato válido, descarta a data (vazio é melhor que futuro).
    - ano do número: edital numera legitimamente para o ano SEGUINTE
      ("Edital POSLING nº 1/2024" sai em 2023) — só corrige se > BS+1.
    Devolve (data_ato, ano_ato) corrigidos."""
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})$", bs_data or "")
    if not m:
        return data_ato, ano_ato
    bs_iso = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    bs_ano = int(m.group(3))
    ano_da_data = data_ato[:4] if data_ato else ""
    if data_ato and data_ato > bs_iso:
        # ano-1 só quando o ano estava claramente typado (difere do ano do BS);
        # se o ano já é o do BS, o typo está no dia/mês — não dá pra inferir,
        # e empurrar o ato p/ um ano atrás fabricaria dado errado.
        candidatos = [bs_ano] + ([bs_ano - 1] if ano_da_data != str(bs_ano) else [])
        for cand in candidatos:
            novo = f"{cand:04d}{data_ato[4:]}"
            if novo <= bs_iso and _data_valida(novo):
                data_ato = novo
                break
        else:
            data_ato = ""
        if ano_ato == ano_da_data:        # ano do título veio da mesma data typada
            ano_ato = data_ato[:4] if data_ato else str(bs_ano)
    if not (ano_ato.isdigit() and len(ano_ato) == 4):
        # Ano abreviado ("064/17") ou corrompido por typo/OCR ("026/211", só
        # 3 dígitos) — a fonte não tem padrão fixo, então não dá pra saber com
        # segurança qual dígito reconstruir. Usa o ano do próprio Boletim (a
        # única data certa que temos) em vez de arriscar um chute de século.
        ano_ato = str(bs_ano)
    elif int(ano_ato) > bs_ano + 1:
        ano_ato = str(bs_ano)
    return data_ato, ano_ato


def data_iso(dia, mes_nome, ano):
    m = MESES.get(mes_nome.lower().strip())
    if not m:
        return ""
    try:
        return f"{int(ano):04d}-{m:02d}-{int(dia):02d}"
    except ValueError:
        return ""


def link_sei_processo(proc):
    """Link para a pesquisa pública de processos do SEI/UFF."""
    if not proc:
        return ""
    return ("https://sei.uff.br/sei/modulos/pesquisa/"
            "md_pesq_processo_pesquisar.php?acao_externa=protocolo_pesquisar"
            "&acao_origem_externa=protocolo_pesquisar"
            "&id_orgao_acesso_externo=0&txtProtocoloPesquisa=" + proc)


def link_sei_documento(cod):
    """Link para conferência pública de documento do SEI/UFF (código verificador)."""
    if not cod:
        return ""
    return ("https://sei.uff.br/sei/controlador_externo.php?"
            "acao=documento_conferir&id_orgao_acesso_externo=0&id_documento=" + str(cod))


# --------------------------------------------------------------------------- #
# Parsing de um PDF
# --------------------------------------------------------------------------- #

def metadados_bs(texto_paginas):
    """Descobre número/data/ano do Boletim a partir da capa."""
    bs_num = bs_data = bs_ano = ""
    capa = "\n".join(texto_paginas[:3])
    ma = re.search(r"ANO\s+([IVXLCDM]+)", capa)
    if ma:
        bs_ano = ma.group(1)
    mn = re.search(r"N[.\s]*[°ºo]?\s*(\d{1,4})", capa)
    if mn:
        bs_num = mn.group(1)
    md = DATA_BS_RE.search(capa)
    if md:
        bs_data = md.group(1)
    return bs_num, bs_data, bs_ano


def contexto_secao_pagina(trecho):
    """Extrai seção e página do(s) cabeçalho(s) presentes no trecho do ato."""
    secao = pagina = ""
    m = SECAO_RE.search(trecho)
    if m:
        secao, pagina = m.group(1), m.group(2)
    return secao, pagina


def classifica_verbo(trecho_anterior):
    """Dado o texto que antecede uma referência, retorna o rótulo da relação."""
    melhor = None
    melhor_pos = -1
    for rotulo, rgx in VERBOS_RELACAO:
        for vm in rgx.finditer(trecho_anterior):
            if vm.start() > melhor_pos:      # verbo mais próximo da referência
                melhor_pos = vm.start()
                melhor = rotulo
    return melhor


def detecta_relacoes(ementa, corpo, sigla_atual, numero_atual):
    """
    Encontra relações entre atos. ANCORA na referência a um ato citado e
    procura, nos ~75 caracteres anteriores, o verbo que define a relação
    (Altera/Revoga/Substitui/...). Evita autorreferências e citações de leis.
    """
    relacoes = []
    texto = limpar(ementa) + " ¶ " + limpar(corpo)
    vistos = set()

    for rm in REF_RE.finditer(texto):
        ato_citado = monta_ref(rm)
        if not ato_citado:
            continue
        # ignora autorreferência (mesma sigla + mesmo número)
        num_ref = re.sub(r"\D", "", rm.group("numero") or rm.group("numero2") or "")
        if num_ref and num_ref == re.sub(r"\D", "", numero_atual) and \
           (sigla_atual and sigla_atual.lower() in ato_citado.lower()):
            continue

        anterior = texto[max(0, rm.start() - 75): rm.start()]
        rotulo = classifica_verbo(anterior)
        if not rotulo:
            continue  # referência sem verbo de relação -> ruído (não registra)

        # BS de origem do ato citado (logo após a referência)
        seguinte = texto[rm.end(): rm.end() + 90]
        bs_origem = ""
        bm = BS_REF_RE.search(seguinte)
        if bm:
            bs_origem = f"BS {bm.group(1)} de {bm.group(2)}"

        chave = (rotulo, re.sub(r"\s+", "", ato_citado.lower()))
        if chave in vistos:
            continue
        vistos.add(chave)
        relacoes.append({
            "relacao": rotulo,
            "ato_citado": ato_citado,
            "bs_origem": bs_origem,
            "trecho": limpar(texto[max(0, rm.start() - 75): rm.end() + 30]),
        })
    return relacoes


ACRONIMOS = {"Dts": "DTS", "Rdd": "RDD", "In": "IN", "Ns": "NS", "Os": "OS",
             "Uff": "UFF", "Cuv": "CUV", "Cepex": "CEPEX", "Cepx": "CEPEx"}


def monta_ref(rm):
    tipo = limpar(rm.group("tipo")).title()
    tipo = ACRONIMOS.get(tipo, tipo)
    orgao = re.sub(r"\s+", " ", limpar(rm.group("orgao") or "")).strip(" /.,-()")
    g = rm.groupdict()
    num = limpar(g.get("numero") or g.get("numero2") or "")
    ano = g.get("ano") or g.get("ano2") or ""
    if not re.search(r"\d", num):
        return ""
    partes = [tipo]
    if orgao:
        partes.append(orgao)
    partes.append("nº " + num + ("/" + ano if ano else ""))
    return " ".join(partes)


def _norm_pdf(s):
    """Normaliza o texto extraído do PDF, ANTES de qualquer parsing:
      - NFC: junta diacrítico combinante ('c'+cedilha -> 'ç') que truncava nomes
        ("Picanço"->"Picanc"), pois o regex parava na letra-base.
      - remove hífen-suave e caracteres de largura-zero que o PDF insere.
      - cola apóstrofo de sobrenome com espaço espúrio ("Sant’ Anna"->"Sant’Anna")."""
    s = unicodedata.normalize("NFC", s)
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", s)  # chars de controle (glitch de glifo)
    s = re.sub("[­​‌‍⁠﻿]", "", s)  # hífen-suave / largura-zero
    s = s.replace("\xa0", " ")     # espaço inquebrável -> espaço comum
    s = re.sub(r"([A-Za-zÀ-ÿ]['’])[ \t]+(?=[A-Za-zÀ-ÿ])", r"\1", s)
    return s


def parse_pdf(caminho):
    doc = fitz.open(caminho)
    paginas = [_norm_pdf(doc[p].get_text()) for p in range(doc.page_count)]
    doc.close()

    bs_num, bs_data, bs_ano = metadados_bs(paginas)
    arquivo = os.path.basename(caminho)

    # texto com marcadores de página para localizar seção/página por offset
    full = ""
    page_offsets = []  # (offset_inicio, numero_pagina, texto_pagina)
    for i, t in enumerate(paginas):
        page_offsets.append((len(full), i + 1, t))
        full += t + "\n"

    # Localiza todos os títulos (pulando o SUMÁRIO, onde aparecem só listas).
    # Junta o regex principal (TIPO em CAIXA ALTA) com o das portarias SIGA em
    # Title Case (ancoradas no marcador "UFFPOR...") e o de títulos curtos sem
    # data por extenso (ancorados no separador "# # # # # #"), ordena por
    # posição e descarta sobreposições (um SIGA em CAIXA ALTA casa nos dois).
    # BOUNDARY_NAO_ATO_RE entra na MESMA lista (define onde o ato anterior
    # termina) mas é filtrado abaixo antes de virar um "ato" de saída.
    brutos = sorted(list(TITULO_RE.finditer(full)) + list(TITULO_SIGA_RE.finditer(full))
                     + list(TITULO_CURTO_RE.finditer(full)) + list(BOUNDARY_NAO_ATO_RE.finditer(full)),
                    key=lambda m: m.start())
    titulos = []
    for m in brutos:
        # ignora ocorrências dentro do sumário (heurística: sem "RESOLVE"/"Art."
        # logo depois e com muitos títulos colados não é o caso; melhor filtrar
        # por presença do cabeçalho do BS antes). Mantemos todos e filtramos
        # quando o corpo for vazio.
        if titulos and m.start() < titulos[-1].end():
            continue                      # sobreposto ao título anterior
        titulos.append(m)

    atos = []
    for idx, m in enumerate(titulos):
        if m.re is BOUNDARY_NAO_ATO_RE:
            continue                      # só fronteira: já cortou o ato anterior, não vira ato
        ini = m.start()
        fim = titulos[idx + 1].start() if idx + 1 < len(titulos) else len(full)
        trecho = full[ini:fim]

        # corpo sem o cabeçalho/rodapé repetido do BS
        corpo = HEADER_BS_RE.sub(" ", trecho)
        corpo = re.sub(_HASH_SEP, " ", corpo)

        # seção/página: procura no trecho IMEDIATAMENTE antes do título também
        ctx = full[max(0, ini - 400): ini + 200]
        secao, pagina = contexto_secao_pagina(ctx)

        tipo = canon_tipo(limpar(m.group("tipo")))
        orgao = limpar(m.group("orgao") or "")
        orgao = re.sub(r"\s+", " ", orgao).strip(" /.,-()")
        numero = limpar(m.group("numero"))
        data_ato = data_iso(m.group("dia"), m.group("mes"), m.group("ano"))
        ano_ato = m.group("ano")
        data_ato, ano_ato = corrige_ano_futuro(data_ato, ano_ato, bs_data)

        # Ementa: texto entre o fim do título e o "RESOLVE"/"O REITOR"/etc.
        pos_resto = m.end()
        resto = full[pos_resto: pos_resto + 1200]
        resto = HEADER_BS_RE.sub(" ", resto)
        ementa = extrai_ementa(resto)

        # Assinante: última linha em CAIXA ALTA antes de cargo/separador
        signatario = extrai_signatario(trecho)

        # Processos SEI e código do documento
        procs = []
        for pm in PROC_RE.finditer(trecho):
            p = normaliza_proc(pm.group(0))
            if p not in procs:
                procs.append(p)
        sei_doc = ""
        sd = SEI_DOC_RE.search(trecho)
        if sd:
            sei_doc = sd.group(1)
        else:
            # padrão "(3442574)" perto da assinatura
            sp = SEI_DOC_PAREN_RE.search(trecho)
            if sp:
                sei_doc = sp.group(1)

        # ID do órgão/sigla "limpo": tira traço/CONJUNTA/conector inicial e /UFF
        sigla = limpa_sigla(orgao)

        relacoes = detecta_relacoes(ementa, corpo, sigla, numero)

        ma = ACAO_EMENTA_RE.search(ementa)
        tipo_acao = ma.group(1).title() if ma else ""

        # Texto do corpo para busca por NOME (pega nomes em tabelas, listas etc.)
        # e SIAPEs explícitas para exibição. O texto cobre o que a ementa não tem.
        corpo_busca = limpar(corpo).lower()[:7000]
        siapes = sorted(set(SIAPE_RE.findall(trecho)))

        # Ementa inferida: só quando NÃO há ementa formal. Resume o dispositivo.
        # "Ementa" que é só o preâmbulo ("FULANA, no uso de suas atribuições,")
        # ou só a cláusula de vigência ("Esta DTS entrará em vigor...") conta
        # como ausente — o que interessa é o que vem depois do RESOLVE:.
        if PREAMBULO_RE.search(ementa) or BOILERPLATE_EMENTA_RE.match(ementa):
            ementa = ""
        ementa_resumo, ementa_inferida = "", False
        if len(ementa.strip()) < 12:
            ementa_resumo, ementa_inferida = sintetiza_ementa(corpo)

        ato = {
            "arquivo": arquivo,
            "bs_numero": bs_num,
            "bs_data": bs_data,
            "secao": secao,
            "pagina": pagina,
            "tipo": tipo,
            "orgao": orgao,
            "sigla": sigla,
            "numero": numero,
            "ano": ano_ato,
            "data_ato": data_ato,
            "identificador": monta_identificador(tipo, sigla, numero, ano_ato),
            "tipo_acao": tipo_acao,
            "ementa": ementa,
            "ementa_resumo": ementa_resumo,
            "ementa_inferida": ementa_inferida,
            "signatario": signatario,
            "processos_sei": procs,
            "processo_sei_principal": procs[0] if procs else "",
            "sei_documento": sei_doc,
            "link_sei_processo": link_sei_processo(procs[0] if procs else ""),
            "link_sei_documento": link_sei_documento(sei_doc),
            "relacoes": relacoes,
            "altera": "; ".join(r["ato_citado"] for r in relacoes if r["relacao"] == "ALTERA"),
            "revoga": "; ".join(r["ato_citado"] for r in relacoes if r["relacao"] in ("REVOGA", "TORNA SEM EFEITO", "ANULA")),
            "substitui": "; ".join(r["ato_citado"] for r in relacoes if r["relacao"] in ("SUBSTITUI", "RETIFICA", "REPUBLICA")),
            "cita": "; ".join(r["ato_citado"] for r in relacoes if r["relacao"] == "CITA"),
            "siapes": siapes,
            "pessoas": extrai_pessoas(trecho),
            "funcoes": extrai_funcoes(trecho, data_ato),
            "aposentadoria": extrai_aposentadoria(trecho),
            "deslocamento": extrai_deslocamento(trecho),
            "corpo_busca": corpo_busca,
        }
        # filtra falsos positivos: títulos capturados dentro do sumário costumam
        # ter corpo muito curto e nenhum verbo "RESOLVE/Art./O ... DA UFF".
        corpo_baixo = trecho.lower()
        tem_corpo = ("resolve" in corpo_baixo or "art." in corpo_baixo
                     or "art " in corpo_baixo or "considerando" in corpo_baixo
                     or len(trecho) > 700)
        if tem_corpo:
            atos.append(ato)

    return atos, {"bs_numero": bs_num, "bs_data": bs_data, "bs_ano": bs_ano,
                  "arquivo": arquivo, "paginas": len(paginas)}


def monta_identificador(tipo, sigla, numero, ano):
    siglas_curtas = {
        "DETERMINAÇÃO DE SERVIÇO": "DTS",
        "RESUMO DE DESPACHOS E DECISÕES": "RDD",
        "RESUMO DE DESPACHOS": "RDD",
        "INSTRUÇÃO NORMATIVA": "IN",
        "NORMA DE SERVIÇO": "NS",
        "ORDEM DE SERVIÇO": "OS",
    }
    base = siglas_curtas.get(tipo, tipo.title())
    partes = [base]
    if sigla:
        partes.append(sigla)
    partes.append(f"nº {numero}/{ano}")
    return " ".join(partes)


# "Ementa" que na verdade é o PREÂMBULO do ato: a autoridade + "no uso de suas
# atribuições" logo no início (sem frase de ementa antes). Acontece quando o
# corte por autoridade não pega a variação de grafia — melhor detectar e tratar
# como sem-ementa (inferindo do dispositivo) do que exibir o preâmbulo na ficha.
PREAMBULO_RE = re.compile(
    r"(?i)^[^.;]{0,80}\bno uso d(?:e suas?\b|e atribui\w*|as?\s+(?:atribui|compet)\w*)")

# "Ementa" que é só a cláusula de vigência/fecho do ato ("Esta DTS entrará em
# vigor na data de sua assinatura. DOCUMENTO ASSINADO..."): ato sem ementa cujo
# corpo começou a ser lido no lugar errado — trata como sem-ementa e infere.
BOILERPLATE_EMENTA_RE = re.compile(
    r"(?i)^\s*est[ae]\s+(?:dts|determina\w*(?:\s+de\s+servi[çc]o)?|portaria|"
    r"resolu[çc][ãa]o|instru[çc][ãa]o(?:\s+normativa)?|norma|ordem|decis[ãa]o)\b"
    r"[^.;]{0,40}\bentrar?[áa]?\s+em\s+vigor")


def extrai_ementa(resto):
    """A ementa é o texto descritivo logo após o título, antes do dispositivo."""
    # corta no primeiro marcador de início de dispositivo
    # A autoridade cobre masculino E feminino ("A PRÓ-REITORA..." começava o
    # preâmbulo sem corte e virava "ementa" — visto na DTS PROAD 15/2020).
    # [-\s]* tolera "PRÓ- REITOR" (hífen+espaço da quebra de linha do PDF).
    cortes = [r"\bRESOLVE\b", r"\bRESOLVEM\b",
              r"\bR\s+E\s+S\s+O\s+L\s+V\s+E\b",     # "R E S O L V E" espaçado
              r"\b[OA]\s+(?:VICE[-\s]*|PR[ÓO][-\s]*)?REITORA?\b",
              r"\bO\s+CONSELHO\b", r"\b[OA]\s+COORDENADORA?\b",
              r"\b[OA]\s+CHEFE\b", r"\b[OA]\s+DIRETORA?\b",
              r"\b[OA]\s+SUPERINTENDENTE\b", r"\b[OA]\s+PRESIDENTE\b",
              r"\b[OA]\s+DECANA?\b", r"\b[OA]\s+GERENTE\b",
              r"\bConsiderando\b", r"\bArt\.?\s*1"]
    pos = len(resto)
    for c in cortes:
        mm = re.search(c, resto)
        if mm and mm.start() < pos:
            pos = mm.start()
    ementa = limpar(resto[:pos])
    # remove números de página/seção e cabeçalho/rodapé do BS que vazam na ementa
    ementa = re.sub(r"(?i)ANO\s+[IVXLCDM]+\s*[–—-]?\s*N[.°º\s]*\d+", "", ementa)
    ementa = re.sub(r"(?i)SE[ÇC][ÃA]O\s+[IVX]+\s*P[ÁA]?G?[.\s]*\d+", "", ementa)
    ementa = re.sub(r"(?i)P[ÁA]G[.\s]*\d+", "", ementa)
    ementa = re.sub(r"\d{2}/\d{2}/\d{4}", "", ementa)
    # Cabeçalho de página: os pares completos somem; MAS menção à universidade
    # no meio da frase vira "UFF" em vez de sumir — apagar deixava buracos
    # agramaticais ("celebrado entre a e a empresa", "Estabelecer, na , o
    # calendário"): 1.682 ementas afetadas só em 2023-2026.
    ementa = re.sub(r"(?i)MINIST[ÉE]RIO\s+DA\s+EDUCA[ÇC][ÃA]O"
                    r"(?:\s+UNIVERSIDADE\s+FEDERAL\s+FLUMINENSE)?", "", ementa)
    ementa = re.sub(r"(?i)UNIVERSIDADE\s+FEDERAL\s+FLUMINENSE"
                    r"\s*[–—-]+\s*BOLETIM\s+DE\s+SERVI[ÇC]O", "", ementa)
    ementa = re.sub(r"(?i)UNIVERSIDADE\s+FEDERAL\s+FLUMINENSE", "UFF", ementa)
    ementa = re.sub(r"(?i)BOLETIM\s+DE\s+SERVI[ÇC]O", "BS", ementa)
    ementa = re.sub(r"#(?:\s*#)+", "", ementa)  # marcadores "# # # #" de rodapé
    # rodapé de assinatura digital que vaza para dentro da ementa
    ementa = re.sub(r"(?i)\s*documento\s+assinado\s+eletronicamente\b.*$", "", ementa)
    # tira pontuação/rótulo iniciais (". ", ", ", "Ementa:") — ruído muito comum
    ementa = limpar(ementa)
    ementa = re.sub(r"^[\s.,;:–\- ]+", "", ementa)
    ementa = re.sub(r"(?i)^ementa\s*:?\s*", "", ementa)
    # enumerador que sobra na frente ("II - Designar...", "1. No item 3.1...")
    ementa = re.sub(r"^(?:[IVX]{1,4}|\d{1,2})\s*[-–—.)]\s+", "", ementa)
    ementa = re.sub(r"^[\s.,;:–\- ]+", "", ementa)
    return limpar(ementa)[:600]


# --------------------------------------------------------------------------- #
# Ementa INFERIDA: para atos sem ementa formal, resume o próprio dispositivo
# (o texto após "resolve:") em 3ª pessoa. NÃO inventa nada — usa as palavras do
# ato. O resultado é marcado como inferido (ementa_inferida=True) para o portal
# exibir como "resumo automático", nunca confundido com a ementa oficial.
# --------------------------------------------------------------------------- #
# Verbo do dispositivo (infinitivo) -> forma na 3ª pessoa do singular.
_VERBO_EMENTA = [
    (r"tornar\s+sem\s+efeito", "Torna sem efeito"), (r"tornar\s+p[úu]blico", "Torna público"),
    (r"designar", "Designa"), (r"nomear", "Nomeia"), (r"exonerar", "Exonera"),
    (r"dispensar", "Dispensa"), (r"constituir", "Constitui"), (r"instituir", "Institui"),
    (r"conceder", "Concede"), (r"autorizar", "Autoriza"), (r"prorrogar", "Prorroga"),
    (r"alterar", "Altera"), (r"revogar", "Revoga"), (r"aprovar", "Aprova"),
    (r"homologar", "Homologa"), (r"retificar", "Retifica"), (r"republicar", "Republica"),
    (r"delegar", "Delega"), (r"criar", "Cria"), (r"estabelecer", "Estabelece"),
    (r"remover", "Remove"), (r"redistribuir", "Redistribui"), (r"lotar", "Lota"),
    (r"cancelar", "Cancela"), (r"suspender", "Suspende"), (r"determinar", "Determina"),
    # "declarar vago" ANTES de "declarar": o verbo-frase leva o objeto junto
    # ("Declara vago o cargo de..."), senão o corte em "a partir de" deixava
    # só "Declara vago." seco (266 casos em 2023-2026).
    (r"fixar", "Fixa"), (r"declarar\s+vagos", "Declara vagos"),
    (r"declarar\s+vago", "Declara vago"), (r"declarar", "Declara"),
    (r"convalidar", "Convalida"),
    (r"reconduzir", "Reconduz"), (r"destituir", "Destitui"), (r"substituir", "Substitui"),
    (r"conceituar", "Conceitua"), (r"ratificar", "Ratifica"), (r"dispor", "Dispõe sobre"),
    (r"interromper", "Interrompe"), (r"aposentar", "Aposenta"), (r"aplicar", "Aplica"),
    (r"atribuir", "Atribui"), (r"extinguir", "Extingue"), (r"atualizar", "Atualiza"),
    (r"publicar", "Publica"), (r"readaptar", "Readapta"), (r"incluir", "Inclui"),
    (r"excluir", "Exclui"), (r"abonar", "Abona"), (r"conferir", "Confere"),
    (r"transferir", "Transfere"), (r"anular", "Anula"), (r"validar", "Valida"),
]
_ENUM_EMENTA_RE = re.compile(
    r"(?i)^\s*(?:art\.?\s*\d+[ºo°.\-]*\s*[-–]?\s*|[ivx]{1,4}\s*[-–.)]\s*|"
    r"\d+\s*[-–.)]\s*|[a-z]\s*[-)]\s*|§\s*\d*[ºo°]?\s*|par[áa]grafo\s+\S+\s*[-–.:]?\s*)")
# Corte do objeto: ";" ou "." de fim de frase (NÃO o "." interno de nº de
# processo "23069.002753"), próximo item (II-, IV-), ou início de cláusula acessória.
_STOP_EMENTA_RE = re.compile(
    r"(?i)(?:;|\.(?=\s|$)|\bII+\s*[-–]|\bIV\s*[-–]|,?\s*\bmatr[íi]cula\b|,?\s*\bsiape\b|"
    r",?\s*\bc[óo]digo\b|\ba partir\b|\bcom valid|\bpelo per[íi]odo\b|\bno per[íi]odo\b|"
    r"\btendo em vista\b|\bem virtude\b|\bnos termos\b|\bem substitui|"
    r"\bpublique-se\b|\bregistre-se\b|\bfica\b)")
# Cláusula acessória inicial a descartar para chegar ao objeto real. Ex.:
# "dispensar, A PEDIDO, A PARTIR DE 04/08/2025, Fulano..." -> "Fulano...".
_CLAUSULA_INI_RE = re.compile(
    r"(?i)^\s*,?\s*(?:"
    r"consoante\b[^,]*|conforme\b[^,]*|nos termos\b[^,]*|tendo em vista\b[^,]*|"
    r"de acordo com\b[^,]*|com base n[oa]\b[^,]*|a pedido|"
    r"a partir d[eo]\b[^,]*|a contar d[eo]\b[^,]*|com efeitos?\b[^,]*|"
    r"pelo per[íi]odo\b[^,]*|no per[íi]odo\b[^,]*|retroativ[oa]\b[^,]*|"
    r"em car[áa]ter\b[^,]*|por \d+[^,]*"
    r")\s*,\s*")
_MINUSC_FRASE = {"de", "da", "do", "das", "dos", "e", "a", "o", "as", "os", "em", "no",
                 "na", "nos", "nas", "para", "por", "com", "sem", "ao", "aos", "à", "às",
                 "que", "sob", "sobre", "entre"}


def _titlecase_frase(s):
    """Title-case de uma frase (só p/ objetos que vieram em CAIXA ALTA): mantém
    conectores minúsculos, capitaliza as demais palavras."""
    out = []
    for i, w in enumerate(s.split()):
        out.append(w.lower() if (i > 0 and _fold(w) in _MINUSC_FRASE)
                   else (w[:1].upper() + w[1:].lower()))
    return " ".join(out)


def sintetiza_ementa(corpo):
    """Resume o dispositivo de um ato sem ementa formal, em 3ª pessoa, usando as
    próprias palavras do ato. Ex.: '...resolve: exonerar Fulano do cargo...' ->
    'Exonera Fulano.'. Retorna (resumo, True) se achou um verbo de dispositivo;
    ('', False) se não deu para inferir (vai p/ o fallback de LLM)."""
    txt = limpar(corpo)
    # "R E S O L V E" com letras espaçadas (diagramação antiga) vira RESOLVE,
    # senão o marcador do dispositivo não é achado e nada é inferido.
    txt = re.sub(r"(?i)\bR\s+E\s+S\s+O\s+L\s+V\s+E(\s+M)?\b",
                 lambda m: "RESOLVEM" if m.group(1) else "RESOLVE", txt)
    m = re.search(r"(?i)\bresolve[m]?\b\s*:?\s*", txt)
    disp = txt[m.end():] if m else txt
    for _ in range(4):                       # pula "I -", "Art. 1º -", "a)"...
        e = _ENUM_EMENTA_RE.match(disp)
        if not e or e.end() == 0:
            break
        disp = disp[e.end():]
    disp = disp.lstrip(" -–.)")
    verbo3 = resto = None
    for pat, t3 in _VERBO_EMENTA:
        mm = re.match(r"(?i)" + pat + r"\b", disp)
        if mm:
            verbo3, resto = t3, disp[mm.end():]
            break
    if not verbo3:
        return "", False
    for _ in range(3):                            # pula cláusulas acessórias iniciais
        novo = _CLAUSULA_INI_RE.sub("", resto, count=1)
        if novo == resto:
            break
        resto = novo
    sm = _STOP_EMENTA_RE.search(resto)
    obj = (resto[:sm.start()] if sm else resto).strip(" ,;.:-–")
    obj = re.sub(r"\s+", " ", obj)
    if len(obj) < 2:
        return "", False
    # title-case runs de 2+ palavras em CAIXA ALTA (nomes/benefícios embutidos):
    # "RENATA HELENA MARTO" -> "Renata Helena Marto"; acrônimos isolados ficam.
    obj = re.sub(r"[A-ZÀ-Þ][A-ZÀ-Þ0-9'’\-]*(?:\s+[A-ZÀ-Þ][A-ZÀ-Þ0-9'’\-]*)+",
                 lambda mm: _titlecase_frase(mm.group(0)), obj)
    ementa = re.sub(r"\s+([,;.])", r"\1", f"{verbo3} {obj}").strip(" ,;:-–")
    return limpar(ementa)[:280].rstrip(" ,;:-–") + ".", True


def extrai_signatario(trecho):
    """Heurística: nome em CAIXA ALTA seguido por um cargo conhecido."""
    linhas = [l.strip() for l in trecho.splitlines() if l.strip()]
    cargos = re.compile(r"(REITOR|DIRETOR|COORDENADOR|CHEFE|PRESIDENTE|SUPERINTEND|"
                        r"PR[ÓO]-REITOR|VICE|SECRET|DECAN)", re.I)
    for i, l in enumerate(linhas):
        if cargos.search(l) and i > 0:
            cand = linhas[i - 1]
            letras = re.sub(r"[^A-Za-zÀ-Ú]", "", cand)
            if letras and cand.upper() == cand and len(letras) >= 6:
                return limpar(cand)
    return ""


def _fold(s):
    """minúsculas e sem acento, para comparar com a blocklist de nomes."""
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn").lower()


# Conectores "de/do/da" que precedem a sigla no título ("... DO RCM/RIC Nº").
_CONECT_SIGLA = {"do", "da", "de", "dos", "das"}


def limpa_sigla(orgao):
    """Normaliza a sigla do órgão emissor capturada no título do ato.

    O título traz separadores e qualificadores que NÃO fazem parte da sigla:
    um traço solto ("DETERMINAÇÃO DE SERVIÇO - DAP/UFF"), o qualificador
    "CONJUNTA" ("PORTARIA CONJUNTA PROPLAN/DCF") ou um conector inicial
    ("... DO RCM/RIC"). Também remove o "/UFF". Ex.:
      "- DAP/UFF"            -> "DAP"
      "CONJUNTA PROPLAN/DCF" -> "PROPLAN/DCF"
      "DO FFE / ISNF"        -> "FFE/ISNF"
    """
    s = orgao or ""
    s = re.sub(r"[–—]", "-", s)                      # unifica traços
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"(?i)/?\s*\bUFF\b", "", s)           # tira "/UFF" ou "UFF" solto
    # remove "CONJUNTA" e conectores iniciais (repete até estabilizar)
    mudou = True
    while mudou:
        mudou = False
        s = s.strip(" /.,-()")
        mc = re.match(r"(?i)CONJUNTA\b\s*", s)
        if mc:
            s = s[mc.end():]; mudou = True; continue
        # "EDITAL DO PROGRAMA DE GESTÃO ESD Nº 2..." — o título padrão do PGD
        # (teletrabalho) põe o NOME DO PROGRAMA entre o tipo e a sigla, e o
        # regex de título capturava tudo como órgão ("PROGRAMA DE GESTÃO ESD").
        # O programa não é setor; a sigla real vem depois. Medido: 37 atos
        # 2022-2025 com esse defeito. "EDITAL" idem, quando vaza junto
        # ("PROGRAMA DE GESTÃO - EDITAL PROAES" -> "PROAES").
        mg = re.match(r"(?i)(?:NOVO )?PROGRAMA DE GEST[ÃA]O\b\s*(?:\(NGPD\))?\s*-?\s*", s)
        if mg:
            s = s[mg.end():]; mudou = True; continue
        me = re.match(r"(?i)EDITAL\b\s*", s)
        if me:
            s = s[me.end():]; mudou = True; continue
        mp = re.match(r"([A-Za-zÀ-ú]+)\b\s*", s)
        if mp and _fold(mp.group(1)) in _CONECT_SIGLA:
            s = s[mp.end():]; mudou = True
    s = re.sub(r"\s*/\s*", "/", s)                   # "FFE / ISNF" -> "FFE/ISNF"
    return s.strip(" /.,-()")


def _titlecase_nome(s):
    def cap(w):
        w = w.lower() if _fold(w) in _CONNECT else (w[:1].upper() + w[1:].lower())
        # capitaliza após apóstrofo/hífen: "sant'anna"->"Sant'Anna", "dal'magro"->"Dal'Magro"
        return re.sub(r"(['’\-])([a-zà-öø-ÿ])", lambda m: m.group(1) + m.group(2).upper(), w)
    return " ".join(cap(w) for w in s.split())


def _limpa_nome(run):
    """Tira verbo/cargo/conector colado nas pontas; exige nome com 2+ palavras.
    Letra solta na ponta também cai ("SEÇÃO IV P.013" gerava "Iv P")."""
    p = run.split()
    while p and (_fold(p[0]) in _BLOCK_NOME or _fold(p[0]) in _CONNECT or len(p[0]) == 1):
        p.pop(0)
    while p and (_fold(p[-1]) in _BLOCK_NOME or _fold(p[-1]) in _CONNECT or len(p[-1]) == 1):
        p.pop()
    return _titlecase_nome(" ".join(p)) if len(p) >= 2 else ""


_RESOLVE_CORTE = re.compile(r"(?i)\bR\s*E\s*S\s*O\s*L\s*V\s*E")


def nome_antes_siape(texto, pos, ignora=None):
    """Nome mais próximo ANTES da matrícula (posição pos), pulando cargos/verbos.
    `ignora(nome)`, opcional, descarta um candidato e deixa a busca seguir para
    o anterior (p.ex. nome que é pedaço do nome da unidade)."""
    # 230 (era 170): com o corte no RESOLVE segurando o preâmbulo, a janela
    # maior deixa de decapitar nomes longe da matrícula ("Lenin dos Santos
    # Pires (SIAPE..." a 172 chars virava "Santos Pires").
    ini = max(0, pos - 230)
    janela = texto[ini:pos]
    # A janela não pode começar no meio de palavra: um corte dentro de
    # "SUBCHEFE" virava "HEFE", que escapa do _BLOCK_NOME e era devolvido
    # como nome ("Hefe do Departamento de Letras Estrangeiras Modernas").
    if ini > 0 and texto[ini - 1].isalnum():
        janela = re.sub(r"^\S+\s*", "", janela, count=1)
    # O designado vem sempre DEPOIS do "RESOLVE"; antes dele fica o preâmbulo
    # com a autoridade em caixa alta ("O CHEFE DO DEPARTAMENTO..."), que nas
    # DTS ganhava do nome verdadeiro (em title-case) na prioridade de caixa alta.
    corte = None
    for corte in _RESOLVE_CORTE.finditer(janela):
        pass
    if corte:
        janela = janela[corte.end():]
    # Portaria padrão: nomes em CAIXA ALTA, cargos/ocupação em title-case.
    # Tenta primeiro sequências totalmente em maiúsculas para evitar capturar o cargo.
    for m in reversed(list(NOME_CAPS_RE.finditer(janela))):
        nome = _limpa_nome(m.group(0))
        if nome and not (ignora and ignora(nome)):
            return nome
    for m in reversed(list(NOME_RE.finditer(janela))):
        nome = _limpa_nome(m.group(0))
        if nome and not (ignora and ignora(nome)):
            return nome
    return ""


def extrai_pessoas(trecho):
    """[{nome, siape}] das pessoas citadas — uma por matrícula, sem repetir."""
    pessoas, vistos = [], set()
    for m in SIAPE_RE.finditer(trecho):
        s = m.group(1)
        if s in vistos:
            continue
        vistos.add(s)
        pessoas.append({"nome": nome_antes_siape(trecho, m.start()), "siape": s})
    return pessoas


# ===========================================================================
#  CHEFIAS — designações/dispensas de função (Chefe, Coordenador, Diretor...)
#  Extrai só do DISPOSITIVO ("...função de Chefe do Departamento X..."), que é
#  a fonte autoritativa. Cada evento vira uma linha rastreável a um ato; a
#  projeção do "titular atual" (último não substituído) é feita no servidor.
# ===========================================================================
_MINUSC = {"de", "da", "do", "das", "dos", "e", "em", "na", "no", "nas", "nos", "a", "o",
           "ao", "aos", "com", "para", "por", "sob", "sem", "entre", "sobre"}

_HIFEN_QUEBRA = r"[-\s]*"   # tolera "Pró-" no fim de linha + "reitor" na linha seguinte
_PFX_CARGO = r"(?:Vice%s|Sub%s)?" % (_HIFEN_QUEBRA, _HIFEN_QUEBRA)
# Cargos de direção/assessoramento. A lista é branca de propósito: o gatilho
# aceita "cargo de X" sem o "de direção", então qualquer nome aqui também passa
# a casar CARGO EFETIVO. Medido no corpus antes de mexer:
#   Assessor[ae]     — 45 "cargo de direção de" + 14 "cargo de"; os 14 são
#                      assessoria de verdade ("assessor de pesquisa do CES").
#   Prefeito         — 4, o Prefeito Universitário. Corregedor[ae] — 2.
#   Secretário-Geral — 5, só na forma hifenizada.
# FICARAM DE FORA, e não é esquecimento:
#   Secretário (solto) — 8 de direção contra 60 de "para o cargo de secretário
#                      EXECUTIVO, por não apresentar documentação": isso é
#                      eliminação em concurso, cargo efetivo. Entraria como 60
#                      designações falsas.
#   Procurador (solto) — "aposentado no cargo de procurador federal" é efetivo;
#                      ganho de 3 não paga o risco.
# Mesma lógica de professor/assistente/técnico, que também não entram: são o
# emprego da pessoa, não posição de direção.
_NUC_CARGO = (r"(?:Chefes?|Coordenador[ae]?(?:es)?|Diretor[ae]?(?:es)?|Superintendentes?|"
              r"Gerentes?|Decan[oa]s?|Assessor[ae]?(?:es)?|Prefeit[oa]s?|Corregedor[ae]?(?:es)?|"
              r"Secret[áa]ri[oa]%sGeral|Pr[óo]%sReitor[ae]?(?:es)?|Reitor[ae]?(?:es)?)"
              % (_HIFEN_QUEBRA, _HIFEN_QUEBRA))
_CARGO_G = r"(?P<cargo>%s%s)" % (_PFX_CARGO, _NUC_CARGO)
_CONECT_CU = r"(?:d[oae]s?|d')"

# Gatilho do dispositivo. Cobre "função de", "função gratificada de",
# "cargo de", "cargo de direção de", "exercício do cargo de"... (toda chefia
# é função gratificada/CD, então a qualificação pode aparecer no meio).
# O lookbehind exige fronteira de palavra antes da preposição: sem ele o "o"
# colado em "nO cargo de..." casava, e menção DESCRITIVA virava designação
# ("A vaga está vinculada à atuação da professora X no cargo de direção de
# Pró-Reitora da PROPPI" gerava um designar-Pró-Reitor a partir da DTS de
# uma banca de concurso).
_TRIG_FUNC = (r"(?<![A-Za-zÀ-ÿ])(?P<prep>d[ao]s?|para\s+(?:as?|os?)|pel[ao]|a|as|o|os)\s+"
              r"(?:exerc[íi]cio\s+d[ao]\s+)?"
              r"(?:fun[çc](?:[ãa]o|[õo]es)|cargo)\s+"
              r"(?:(?:gratificad|comissionad)[ao]s?\s+|de\s+confian[çc]a\s+|em\s+comiss[ãa]o\s+)?"
              r"de\s+(?:dire[çc][ãa]o\s+de\s+)?")
# A vírgula só encerra a unidade quando NÃO for parte do próprio nome: em
# "Pró-Reitor de Pesquisa, Pós-Graduação e Inovação" a vírgula é interna
# (seguida de Palavra Capitalizada) e truncava a unidade em só "Pesquisa" —
# fragmentando a chave e quebrando o pareamento designação↔exoneração.
# Vírgula seguida de minúscula ("Divisão X, da Superintendência...") continua
# encerrando (o que vem depois é o órgão-pai, não parte do nome).
# O traço opcional antes dos marcadores cobre "... e Inovação - Código CD-2".
# O sufixo de gênero "(a)"/"(A)" colado no cargo ("Pró-Reitor(a) da...",
# "Coordenador(a) do...", comum em portarias de nomeação) impedia o casamento
# do conector logo depois do cargo — a designação inteira era perdida.
FUNCAO_RE = re.compile(
    _TRIG_FUNC + _CARGO_G + r"(?:\s*\([aA]\))?\s+" + _CONECT_CU + r"\s+"
    r"(?P<unidade>[A-ZÀ-Úa-zà-ú(][^;:.]{2,90}?)"
    r"(?=\s*(?:[-–—]\s*)?(?:,(?!\s*(?-i:[A-ZÀ-Ú][a-zà-ú]))|;|\.|:|\bc[óo]digo\b|\bc[óo]d\b|\bs[íi]mbolo\b|\bFG[- ]?\d|\bCD[- ]?\d|"
    r"\bFCC\b|\bFUC\b|\bn[.ºo°]|\ba partir\b|\bpelo per|\bno per[íi]odo\b|\bcom valid|"
    r"\bem substitui|\bda Universidade\b|/UFF|\bem virtude\b|\bdurante\b|$))", re.I)

_TIPO_SO_UNID = {"curso", "departamento", "programa", "instituto", "faculdade", "escola",
                 "divisao", "secao", "setor", "nucleo", "coordenacao", "coordenadoria",
                 "diretoria", "gerencia", "reitoria", "unidade", "polo", "colegiado"}
# "nomear/nomeia" é o par de ENTRADA dos cargos de direção (CD), como
# "designar" é o das funções — e faltava aqui, embora "exoner" (a SAÍDA do CD)
# já estivesse. A assimetria não era inofensiva: numa portaria que exonera um e
# nomeia outro no mesmo texto (padrão comum), a janela achava só o "Exonerar"
# anterior e o NOMEADO entrava como dispensado — erro invertido e silencioso.
#
# O (?!c) é o que separa VERBO de MENÇÃO, e não é detalhe: só o VERBO conta,
# porque é ele o dispositivo. O substantivo "nomeação" aparece em oração
# explicativa dando o MOTIVO de uma dispensa — "dispensar, em virtude de sua
# nomeação para diretor do Centro..." — e como _acao_func fica com o ÚLTIMO
# verbo da janela, casar o substantivo invertia 34 dispensas reais em
# designações (medido). Mesma regra que já vale para aposentadoria: classifique
# pelo dispositivo, não por menção. _fold() tira o acento, então "nomeação"
# chega aqui como "nomeacao" — daí excluir o "c" seguinte, que deixa passar
# nomear/nomeado/nomeada. "nomei" cobre a ementa ("Nomeia").
_VERBO_FUNC = re.compile(r"design|nomea(?!c)|nomei|dispens|exoner|destitu")
_SUBST_FUNC = re.compile(r"(?i)substitut|eventual|pro\s*tempore|respond|interin|exerc[íi]cio eventual")
_ANAFORA_UNID = re.compile(r"\b(referid|mesm|respectiv|citad|aludid|supracitad|present|seguinte|propri)")

# Nomeação de pessoa EXTERNA (convidado/sem vínculo => SEM SIAPE): p.ex. o Reitor
# "Nomear Marina Vieira Gontijo, para exercer como Convidado, o Cargo de
# Superintendente da ...". Como não há matrícula, o nome é o único identificador.
# Só captura com verbo de nomeação + Nome Próprio (capitalizado) + "para exercer"
# logo antes do gatilho do cargo — a exigência de SIAPE é o que segura o ruído
# nos demais casos, então este atalho tem que ser bem restrito.
# Nome próprio em Title Case OU CAIXA ALTA (portarias de nomeação usam os dois):
# token começa com maiúscula e o resto pode ser maiúsculo ("MARINA") ou
# minúsculo ("Marina"); conectores de/da/dos entre os tokens em qualquer caixa.
_NOME_PROP = r"[A-ZÀ-Ú][A-ZÀ-Úa-zà-ú]+(?:\s+(?:d[aeo]s?\s+|D[AEO]S?\s+)?[A-ZÀ-Ú][A-ZÀ-Úa-zà-ú]+){1,4}"
# A vírgula depois do nome é opcional: portarias de convidado externo tanto
# escrevem "Nomear FULANO, para exercer" quanto "Nomear FULANO para exercer"
# (ex.: Vera Cajazeiras, Pró-Reitora de Administração convidada). O "para
# exercer" logo após o nome já é âncora forte o bastante.
_NOMEIA_EXT = re.compile(r"\b[Nn]omear\s+(?P<nome>" + _NOME_PROP + r")\s*,?\s*para\s+exercer\b")
# Palavras que denunciam que o "nome" é na verdade um coletivo/genérico.
_NAO_NOME = re.compile(r"\b(comiss|membro|docente|servidor|professor|grupo|equipe|"
                       r"seguinte|abaixo|relacionad)", re.I)


def _nome_externo_antes(trecho, pos_gatilho):
    """Nome do convidado nomeado, quando o ato não traz SIAPE. Exige que a
    nomeação ('Nomear FULANO, para exercer') termine logo antes do gatilho do
    cargo (até ~50 chars de 'como Convidado,' no meio). Senão retorna ''."""
    jan_ini = max(0, pos_gatilho - 150)
    jan = trecho[jan_ini:pos_gatilho]
    ult = None
    for x in _NOMEIA_EXT.finditer(jan):
        ult = x
    if not ult:
        return ""
    if pos_gatilho - (jan_ini + ult.end()) > 50:   # nome longe do cargo: não é a mesma frase
        return ""
    if _NAO_NOME.search(ult.group("nome")):         # coletivo/genérico, não é pessoa
        return ""
    return _limpa_nome(ult.group("nome"))


# Mandato da designação: PRAZO e DATA DE INÍCIO ----------------------------- #
# A designação de chefia é AUTOLIMITADA — ela traz a própria validade ("com
# mandato de 04 (quatro) anos"). Por isso o Boletim quase nunca publica a
# "revogação" ao fim do mandato: ela seria redundante. A dispensa, quando
# aparece, é o ato de encerrar ANTES da hora (medido no corpus: 83% das
# dispensas saem >90 dias antes do fim do prazo). Consequência prática: o fim
# do mandato só existe como DADO se for calculado daqui — não há ato para ele.
#
# Ancorar em "mandato" é obrigatório, não conveniência: o corpus tem "pelo
# prazo de 03 (três) anos" em LICENÇA para tratar de interesses particulares,
# que não é mandato nenhum. Casar "N (extenso) anos" solto importaria isso como
# se fosse prazo de chefia. Mesmo princípio de intent-anchoring da aba Prazos.
_MANDATO_RE = re.compile(r"mandato\s+de\s+(?P<n>\d{1,2})\s*\(", re.I)
# A unidade vem DEPOIS do extenso entre parênteses ("04 (quatro) anos"), então
# tem que pular o fecha-parênteses antes de ler "anos"/"meses".
_MANDATO_UNID_RE = re.compile(r"^[^)]{0,20}\)\s*(?P<unid>m[eê]s(?:es)?|anos?)", re.I)
# Mandato-tampão: quem COMPLETA o mandato do antecessor. O relógio começou com
# o ANTECESSOR, não com este ato — "complementando assim, o mandato de 04
# (quatro) anos, iniciado em 29 de abril de 2003". Somar o prazo à data deste
# ato daria ao substituto um mandato novo em folha, quando ele pode ter só
# meses pela frente. São ~7% das designações com prazo, e justamente os
# substitutos: a população mais propensa a esticar sem que ninguém veja.
_INICIADO_RE = re.compile(
    r"(?:iniciado|com\s+in[íi]cio)\s+em\s+(?P<d>\d{1,2})\s+de\s+(?P<m>\w+)\s+de\s+(?P<a>\d{4})", re.I)
# Início declarado: "Designar, a partir de 30/03/2026, FULANO, ... com mandato
# de 04 (quatro) anos, a função de ...". Sem isso o início vira a data do ato,
# que é só a data em que o BS publicou — não a data em que o mandato corre.
_APARTIR_RE = re.compile(r"a\s+partir\s+de\s+(?P<d>\d{1,2})[./](?P<m>\d{1,2})[./](?P<a>\d{2,4})", re.I)


def _data_dmy(m, ano_ref):
    """DD/MM/AA(AA) -> ISO. Ano de 2 dígitos resolve pelo século do ato."""
    d, mes, a = int(m.group("d")), int(m.group("m")), int(m.group("a"))
    if a < 100:
        a = 2000 + a if ano_ref and 2000 + a <= ano_ref + 1 else 1900 + a
    try:
        return f"{a:04d}-{mes:02d}-{d:02d}" if date(a, mes, d) else ""
    except ValueError:
        return ""


def extrai_mandato(trecho, pos_cargo, data_ato):
    """(prazo_meses|None, data_inicio, origem) do mandato da designação.

    origem: 'tampao'    -> completa mandato do antecessor (início é o dele)
            'declarado' -> "a partir de DD/MM/AAAA" no próprio ato
            'data_ato'  -> nada declarado; usa a data do ato (aproximação)

    Janela ancorada no cargo: "com mandato de" e "a partir de" vêm ANTES do
    gatilho do cargo; o "iniciado em" do tampão vem DEPOIS da unidade. Buscar
    no ato inteiro pegaria prazo de um artigo e cargo de outro nos atos que
    designam várias pessoas.
    """
    ini, fim = max(0, pos_cargo - 400), pos_cargo + 300
    jan = trecho[ini:fim]
    ano_ref = int(data_ato[:4]) if data_ato[:4].isdigit() else 0

    m = _MANDATO_RE.search(jan)
    prazo = None
    if m:
        n = int(m.group("n"))
        u = _MANDATO_UNID_RE.match(jan[m.end():])
        # "mandato de 04 (quatro)" sem a palavra "anos" existe no legado; o
        # padrão da casa é ano, então é o default quando a unidade não vem.
        prazo = n if (u and u.group("unid").lower().startswith("m")) else n * 12

    inicio, origem = data_ato, "data_ato"
    if m:
        t = _INICIADO_RE.search(jan[m.start():])
        if t:
            d = data_iso(t.group("d"), t.group("m"), t.group("a"))
            if d:
                inicio, origem = d, "tampao"
    if origem == "data_ato":
        p = _APARTIR_RE.search(jan[:pos_cargo - ini])
        if p:
            d = _data_dmy(p, ano_ref)
            if d:
                inicio, origem = d, "declarado"

    # Guarda de sanidade: início muito longe do ato é erro de captura (OCR na
    # data, "a partir de" de outra coisa). Tampão olha até 6 anos para trás
    # (mandato de 4 anos + folga); nada olha mais que 2 anos à frente.
    if origem != "data_ato" and data_ato:
        if not (_soma_anos(data_ato, -6) <= inicio <= _soma_anos(data_ato, 2)):
            inicio, origem = data_ato, "data_ato"
    return prazo, inicio, origem


def _soma_anos(iso, n):
    """ISO +- n anos, tolerante a 29/02."""
    try:
        a, m, d = (int(x) for x in iso[:10].split("-"))
        try:
            return date(a + n, m, d).isoformat()
        except ValueError:
            return date(a + n, m, 28).isoformat()
    except Exception:
        return iso


def canon_cargo(c):
    """Normaliza a grafia do cargo, preservando Vice-/Sub."""
    low = _fold(c)
    pref = ""
    m = re.match(r"(vice|sub)[-\s]?", low)
    if m:
        pref = "Vice-" if m.group(1) == "vice" else "Sub"
        low = low[m.end():]
    if "reitor" in low:        base = "Pró-Reitor" if "pro" in _fold(c) else "Reitor"
    elif "superintend" in low: base = "Superintendente"
    elif "coordena" in low:    base = "Coordenador"
    elif "chef" in low:        base = "Chefe"
    elif "dire" in low:        base = "Diretor"
    elif "geren" in low:       base = "Gerente"
    elif "decan" in low:       base = "Decano"
    else:                      base = c.title()
    return ("Sub" + base.lower()) if pref == "Sub" else (pref + base)


def _titulo_unidade(u):
    """Title-case da unidade: conectores minúsculos, siglas curtas preservadas."""
    u = re.sub(r"\s+", " ", u).strip()
    saida = []
    for w in u.split(" "):
        sub = []
        for part in w.split("-"):
            if not part:
                sub.append(part); continue
            f = _fold(part)
            if f in _MINUSC and saida:
                sub.append(part.lower())
            elif part.isupper() and len(part) <= 6 and not re.search(r"[À-Ú]", part):
                sub.append(part)
            else:
                sub.append(part[:1].upper() + part[1:].lower())
        saida.append("-".join(sub))
    return " ".join(saida)


def _limpa_unid(u):
    u = unicodedata.normalize("NFKC", u)                      # desfaz ligaduras (ﬁ->fi)
    u = re.sub(r"\s+", " ", u).strip(" /.,-–·()")
    u = re.sub(r"(?i)\s+d[aoe]st?[ae]?\s+Universidade.*$", "", u)
    u = re.sub(r"(?i)\s+d[ao]\s+UFF\b.*$", "", u)
    u = re.sub(r"\s*\([^)]*$", "", u)                         # parêntese aberto no fim
    u = re.sub(r"(?i)\b(\w+)(\s+\1\b)+", r"\1", u)            # colapsa palavra repetida (em em)
    return u.strip(" /.,-–·")


def _unid_ok(u):
    f = _fold(u)
    return len(f) >= 4 and f not in _TIPO_SO_UNID and not _ANAFORA_UNID.search(f)


def chave_unidade(u):
    """Chave estável p/ casar a mesma unidade escrita de formas diferentes."""
    f = _fold(unicodedata.normalize("NFKC", u))
    f = re.sub(r"[\-–]\s*[a-z]{2,6}$", "", f)                 # sigla final "- gcm"
    f = re.sub(r"\([a-z0-9/]{2,8}\)", "", f)                  # sigla "(pch)"
    f = re.sub(r"\bpos\s*-?\s*graduacao\b", "posgraduacao", f)
    f = re.sub(r"[^a-z0-9]+", " ", f)
    return re.sub(r"\s+", " ", f).strip()


def _acao_func(trecho, pos, prep):
    jan = _fold(trecho[max(0, pos - 300):pos])
    last = None
    for mm in _VERBO_FUNC.finditer(jan):
        last = mm.group(0)
    if last in ("dispens", "exoner", "destitu"):
        return "dispensar"
    if last in ("design", "nomea", "nomei"):
        return "designar"
    return "dispensar" if re.match(r"d[ao]s?$", _fold(prep).strip()) else "designar"


def _pessoa_antes(trecho, pos, ignora=None):
    achou = None
    for m in SIAPE_RE.finditer(trecho[:pos]):
        achou = m
    if achou and pos - achou.start() < 220:
        return achou.group(1), nome_antes_siape(trecho, achou.start(), ignora)
    return "", nome_antes_siape(trecho, pos, ignora)


def extrai_funcoes(trecho, data_ato=""):
    """[{acao, cargo, unidade, unidade_chave, nome, siape, prazo_meses,
    data_inicio, inicio_origem}] — designações/dispensas de chefia/coordenação/
    direção citadas no dispositivo. Sempre exige SIAPE.

    O mandato (prazo/início) só faz sentido na DESIGNAÇÃO — a dispensa encerra
    o mandato de outro ato, não abre um.
    """
    ev, vistos = [], set()
    for m in FUNCAO_RE.finditer(trecho):
        if _SUBST_FUNC.search(trecho[max(0, m.start() - 50):m.start("cargo")]):
            continue
        unid = _limpa_unid(m.group("unidade"))
        if not _unid_ok(unid):
            continue
        # Candidato a nome que é pedaço do nome da unidade é lixo ("Mica Ii"
        # em "...Disciplina de MICA II"): rejeitar DURANTE a busca deixa o
        # fallback achar o nome verdadeiro, em vez de só zerar no fim.
        lixo_unid = lambda n: _fold(n) in _fold(unid)
        siape, nome = _pessoa_antes(trecho, m.start("cargo"), lixo_unid)
        if not siape:
            # Sem SIAPE: só segue se for nomeação de convidado externo (nome próprio).
            nome = _nome_externo_antes(trecho, m.start())
            if not nome:
                continue
        if nome and _fold(nome) in _fold(unid):              # nome-lixo da própria unidade
            nome = ""
        cargo = canon_cargo(m.group("cargo"))
        chave = chave_unidade(unid)
        k = (_acao_func(trecho, m.start(), m.group("prep")), cargo.lower(), chave, siape)
        if k in vistos:
            continue
        vistos.add(k)
        reg = {"acao": k[0], "cargo": cargo, "unidade": _titulo_unidade(unid),
               "unidade_chave": chave, "nome": nome, "siape": siape,
               "prazo_meses": None, "data_inicio": "", "inicio_origem": ""}
        if k[0] == "designar":
            reg["prazo_meses"], reg["data_inicio"], reg["inicio_origem"] = \
                extrai_mandato(trecho, m.start("cargo"), data_ato)
        ev.append(reg)
    return ev


# Aposentadorias: classifica pelo DISPOSITIVO de concessão, nunca por menção
# solta — desde 2023 é comum a Portaria de nomeação/vacância dizer "a vacância
# corresponde à aposentadoria voluntária de fulano, publicada pela Portaria
# nº X" (retrospecto da vaga, não uma concessão nova); contar isso dobra o
# número e não tem nada a ver com o ato em questão.
# O RÓTULO mudou de forma ao longo dos anos: "Concede aposentadoria
# compulsória" (fraseado recente) e "Declara aposentado(a), compulsoriamente"
# (legado 2014-2015, verbo "declarar" + advérbio, não "compulsória" como
# adjetivo) são o MESMO ato — sem os dois padrões a compulsória do legado
# ficava quase invisível (achado real: 2016-2019 caíam a ~0/ano só por causa
# do fraseado antigo, enquanto voluntária continuava na casa de centenas).
# Quando não há rótulo nenhum, cai pra BASE LEGAL: art. 40 da Constituição,
# §1º, inciso I=invalidez / II=compulsória / III=voluntária (Regime Próprio
# de Previdência) — cobre concessões que só citam o dispositivo legal.
# Sem nenhum dos dois sinais: 'Indefinida' (não desaparece — fica visível
# como não-classificada, mesmo espírito dos "ambíguos" em resolver_relacoes).
_APOSENT_RETRO_RE = re.compile(
    r"c[oó]digo\s+de\s+vaga|origem\s+da\s+vaga|decorrente\s+da\s+(?:posse|aposentadoria)|"
    r"oriund[ao]\s+de\s+vac[aâ]ncia|vac[aâ]ncia\s+corresponde|corresponde\s+[aà]\s+aposentadoria",
    re.I,
)
_APOSENT_DISPOSITIVO_RE = re.compile(
    r"conced\w*\s+(?:a\s+)?aposentadoria|declara\w*\s+aposentad[oa](?:\s*\([aA]\))?\b", re.I,
)
_APOSENT_COMPULSORIA_RE = re.compile(r"aposentadoria\s+compuls[oó]ria|compulsoriamente", re.I)
_APOSENT_VOLUNTARIA_RE = re.compile(r"aposentadoria\s+volunt[aá]ria", re.I)
_APOSENT_INVALIDEZ_RE = re.compile(r"aposentadoria\s+por\s+(?:invalidez|incapacidade)", re.I)
_ART40_RE = re.compile(r"art(?:igo)?\.?\s*40\b", re.I)
_INCISO_ART40_RE = re.compile(r"inciso\s+(i{1,3})\b|§\s*1[ºo]?[^.;]{0,15}?\b(i{1,3})\b", re.I)
# Retificação que CITA uma concessão anterior: "...a portaria nº X de DD/MM/AAAA,
# publicada no DOU..., QUE concedeu aposentadoria a fulano..." — "que" logo antes
# do verbo é oração relativa (descreve a portaria REFERENCIADA), não o dispositivo
# deste ato (cujo verbo real é "alterar/retificar", visto no início do "resolve:").
# Achado real: sem isso, retificações de fundamentação legal/proporcionalidade
# eram contadas como concessões novas — 34% dos casos no legado 2001-2014,
# 11% em 2015-2022 (a compulsória "quase 0" nunca foi por causa disso, mas a
# poluição existia e inflava as contagens).
_QUE_ANTES_RE = re.compile(r"\bque\s*$", re.I)


def extrai_aposentadoria(trecho):
    """{'tipo': 'Voluntária'|'Compulsória'|'Invalidez'|'Indefinida', 'baseLegal': str}
    ou None se o ato não CONCEDE aposentadoria (mera menção/retrospecto/retificação
    de ato anterior não conta)."""
    pos_disp = [
        m.start() for m in _APOSENT_DISPOSITIVO_RE.finditer(trecho)
        if not _APOSENT_RETRO_RE.search(trecho[max(0, m.start() - 60):m.start() + 60])
        and not _QUE_ANTES_RE.search(trecho[max(0, m.start() - 10):m.start()])
    ]
    if not pos_disp:
        return None
    if _APOSENT_COMPULSORIA_RE.search(trecho):
        return {"tipo": "Compulsória", "baseLegal": ""}
    if _APOSENT_VOLUNTARIA_RE.search(trecho):
        return {"tipo": "Voluntária", "baseLegal": ""}
    if _APOSENT_INVALIDEZ_RE.search(trecho):
        return {"tipo": "Invalidez", "baseLegal": ""}
    for ap in (m.start() for m in _ART40_RE.finditer(trecho)):
        im = _INCISO_ART40_RE.search(trecho[max(0, ap - 100):ap + 100])
        if not im:
            continue
        romano = (im.group(1) or im.group(2) or "").lower()
        if romano == "iii":
            return {"tipo": "Voluntária", "baseLegal": "art. 40, §1º, III, CF"}
        if romano == "ii":
            return {"tipo": "Compulsória", "baseLegal": "art. 40, §1º, II, CF"}
        if romano == "i":
            return {"tipo": "Invalidez", "baseLegal": "art. 40, §1º, I, CF"}
    return {"tipo": "Indefinida", "baseLegal": ""}


# Deslocamento de servidor (Lei 8.112/90): REMOÇÃO = dentro da própria UFF
# (art. 36); REDISTRIBUIÇÃO = cargo entra/sai da UFF p/ outro órgão (art. 37).
_RED_QQ = re.compile(r"redistribu", re.I)
_RED_EXCL = re.compile(
    r"redistribuiç[ãa]o\s+de\s+cargos\s+de\s+direç"        # CD/FG genérico do MEC
    r"|funç[õo]es\s+gratificadas\s+do\s+minist"
    r"|vagas?\s+.{0,30}?redistribu"                         # vaga de edital/bolsa
    r"|redistribu\w+\s+os?\s+dados"                         # IN sobre dados
    r"|redistribu\w+\s+para\s+recredenciamento"
    r"|redistribu\w+\s+.{0,40}?espaç"                       # espaço físico/interno, não servidor
    r"|entende-se\s+por\s+redistribu"                       # definição normativa (IN), não evento
    r"|remo[çc][ãa]o\s*/\s*redistribui"                     # cláusula de atribuição de cargo (rotina do setor)
    r"|redistribu\w+\s+pela\s+portaria\s+mec"               # proveniência da VAGA, não do servidor deste ato
    r"|redistribu\w+\s+por\s+meio\s+da\s+portaria"
    r"|c[óo]digo\s+de\s+vaga.{0,30}?redistribu"
    r"|redistribu\w+\s+(?:através\s+de\s+)?portaria\s+mec\s+n", re.I)
# Instituição de origem/destino EXTERNA (não-UFF). "Ministério da Educação" fica
# de fora de propósito: é o rodapé padrão de TODO ato (a UFF também é MEC), não
# indica origem/destino externo — incluí-lo sem essa exceção derrubou casos reais
# (medido: 9 atos de 2022-2024 viraram falso-negativo quando testado sem o
# `(?!educaç)`, porque o rodapé "Ministério da Educação Universidade Federal
# Fluminense" cai dentro da janela de busca do destino).
_INST_EXTERNA = re.compile(
    r"universidade\s+federal\s+(?!fluminense)\w[\wçãáéíóúâêô]*"
    r"|universidade\s+estadual\s+\w+"
    r"|universidade\s+federal\s+rural\s+\w+"
    r"|fundaç[ãa]o\s+universidade\s+(?:de|do|federal)"
    r"|instituto\s+federal\s+\w+"
    r"|centro\s+federal\s+de\s+educaç[ãa]o"
    r"|cefet\b"
    r"|minist[ée]rio\s+d[aeo]\s+(?!educaç)\w+"
    r"|departamento\s+nacional\s+de\s+\w+"
    r"|instituto\s+do\s+patrim[ôo]nio\s+hist[óo]rico", re.I)
_UFF_MARK = re.compile(r"desta\s+universidade|universidade\s+federal\s+fluminense", re.I)


def classifica_redistribuicao(t):
    """'Entrada', 'Saída' ou None. Não é regex único de propósito: a frase muda
    conforme quem redige (Portaria de lotação: "redistribuído do quadro
    permanente da X, para o desta universidade..."; Decisão do CEPEX: "aprovar
    a redistribuição do docente Y, da X para o Departamento Z"; DTS do DAP:
    só cita a origem, sem "para" — o destino é implícito, é o próprio DAP da
    UFF publicando). O discriminador real é posicional: qual lado de "para" —
    origem ou destino — nomeia uma instituição externa (`_INST_EXTERNA`).
    Medido no corpus 2001-2026 (reprocessamento de 15/07/2026): a versão
    anterior (só "para o desta universidade" / "a UFF por extenso" a até 90
    caracteres de "redistribu") classificava 252 de 1.029 menções não
    excluídas; esta versão classifica 578 (504 Entrada + 74 Saída) — o resto
    é, na amostra revisada, legitimamente outra coisa (proveniência de vaga
    do MEC, redistribuição de espaço físico, cláusula de rotina do setor)."""
    if not _RED_QQ.search(t) or _RED_EXCL.search(t):
        return None
    m = re.search(r"redistribu\w+", t, re.I)
    if not m:
        return None
    janela = t[m.end(): m.end() + 260]
    mp = re.search(r"\bpara\b", janela, re.I)
    if mp:
        antes, depois = janela[:mp.start()], janela[mp.start():mp.start() + 150]
        origem_externa = bool(_INST_EXTERNA.search(antes))
        destino_externo = bool(_INST_EXTERNA.search(depois))
        destino_uff = bool(_UFF_MARK.search(depois)) or (
            re.search(r"\b(?:departamento|instituto|faculdade|escola)\s+de\b", depois, re.I)
            and not destino_externo)
        if origem_externa and not destino_externo:
            return "Entrada"
        if destino_externo and not origem_externa:
            return "Saída"
        if destino_uff and not origem_externa:
            return "Entrada"
        return None
    # sem "para" na janela: padrão DTS do DAP ("redistribuído da/do ORIGEM",
    # sem destino explícito — quem publica o ato já É o destino, a UFF).
    m2 = re.match(r"\s*(?:da|do)\s+(.{0,80})", janela, re.I)
    if m2 and _INST_EXTERNA.search(m2.group(1)):
        return "Entrada"
    return None
# Remoção: "remover" ancorado num servidor/matrícula na MESMA frase (sem ponto).
# Cobre "remover o(a) servidor(a)", "remover, em caráter provisório, a servidora",
# "remover os servidores abaixo". A âncora descarta "remover" de outros contextos.
_REMOVER = re.compile(
    r"\bremover\b[^.]{0,60}?\b(?:servidor|servidora|professor|docente|"
    r"matr[íi]cula\s+siape|siape\s+n)", re.I)
_M_SAUDE = re.compile(r"motivos?\s+de\s+sa[úu]de", re.I)
_M_CONJ = re.compile(r"acompanhar\s+c[ôo]njuge|acompanhar\s+companheir", re.I)
_M_PERMUTA = re.compile(r"permuta", re.I)
_M_OFICIO = re.compile(r"de\s+of[íi]cio|interesse\s+d[ao]\s+administraç", re.I)
_M_PEDIDO = re.compile(r"a\s+pedido", re.I)
# Unidade de DESTINO ("para a UNIDADE - SIGLA[, uorg N]"): prefere a SIGLA (chave
# estável), última ocorrência (a origem vem antes); sem sigla, nome curto.
_DEST_SIGLA = re.compile(
    r"para\s+[oa]s?\s+[^.;:]{3,90}?\s[-–]\s*([a-zà-ú]+(?:/[a-zà-ú]+){0,2})\s*"
    r"(?:[,-–]\s*uorg|\.|:|$)", re.I)
_DEST_NOME = re.compile(
    r"para\s+[oa]s?\s+((?:departamento|instituto|faculdade|escola|coordena\w+|"
    r"divis[ãa]o|superintend\w+|pró-?reitoria|reitoria|hospital|n[úu]cleo|"
    r"setor|se[çc][ãa]o|ger[êe]ncia|diretoria|secretaria|centro)"
    r"[^.;:]{2,55}?)(?:,\s*por\b|,?\s*uorg|\.|$)", re.I)


def extrai_deslocamento(trecho):
    """{'tipo': 'Remoção'|'Redistribuição', 'direcao': 'Interna'|'Entrada'|'Saída',
    'motivo': str, 'setor': str} ou None se o ato não desloca servidor."""
    t = trecho
    direcao_red = classifica_redistribuicao(t)
    if direcao_red:
        return {"tipo": "Redistribuição", "direcao": direcao_red, "motivo": "", "setor": ""}
    if _REMOVER.search(t):
        if _M_SAUDE.search(t): motivo = "Saúde"
        elif _M_CONJ.search(t): motivo = "Acompanhar cônjuge"
        elif _M_PERMUTA.search(t): motivo = "Permuta"
        elif _M_OFICIO.search(t): motivo = "De ofício"
        elif _M_PEDIDO.search(t): motivo = "A pedido"
        else: motivo = ""
        setor = ""
        ms = list(_DEST_SIGLA.finditer(t))
        cand = ms[-1].group(1).upper().strip("/") if ms else ""
        if cand and cand != "UORG" and 2 <= len(cand.replace("/", "")) <= 12:
            setor = cand
        else:
            mn = None
            for mn in _DEST_NOME.finditer(t):
                pass
            if mn:
                setor = re.sub(r"\s+", " ", mn.group(1)).strip().rstrip(" -–").title()[:45]
        return {"tipo": "Remoção", "direcao": "Interna", "motivo": motivo, "setor": setor}
    return None


# --------------------------------------------------------------------------- #
# Saídas
# --------------------------------------------------------------------------- #

COLUNAS = [
    ("identificador", "Identificador"),
    ("tipo", "Tipo"),
    ("sigla", "Órgão/Sigla"),
    ("numero", "Número"),
    ("ano", "Ano"),
    ("data_ato", "Data do ato"),
    ("tipo_acao", "Natureza"),
    ("ementa", "Ementa"),
    ("signatario", "Assinante"),
    ("altera", "Altera"),
    ("revoga", "Revoga / Torna sem efeito"),
    ("substitui", "Substitui / Retifica / Republica"),
    ("cita", "Cita / Fundamenta-se em"),
    ("processo_sei_principal", "Processo SEI"),
    ("sei_documento", "Doc. SEI"),
    ("link_sei_processo", "Link SEI (processo)"),
    ("link_sei_documento", "Link SEI (documento)"),
    ("bs_numero", "Boletim nº"),
    ("bs_data", "Data do Boletim"),
    ("secao", "Seção"),
    ("pagina", "Página"),
    ("arquivo", "Arquivo"),
]


def salvar_csv(atos, caminho):
    import csv
    with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow([c[1] for c in COLUNAS])
        for a in atos:
            w.writerow([a.get(c[0], "") for c in COLUNAS])


def _payload(atos, meta):
    return {"gerado_em": datetime.now().isoformat(timespec="seconds"),
            "boletins": meta, "total": len(atos), "atos": atos}


def salvar_json(atos, meta, caminho):
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(_payload(atos, meta), f, ensure_ascii=False, indent=1)


def salvar_dados_js(atos, meta, caminho):
    """Mesma base em JS para o portal.html funcionar offline (file://)."""
    with open(caminho, "w", encoding="utf-8") as f:
        f.write("window.DADOS = ")
        json.dump(_payload(atos, meta), f, ensure_ascii=False)
        f.write(";")


def salvar_xlsx(atos, caminho):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  (openpyxl não instalado — pulei o XLSX; gere com pip install openpyxl)")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Atos"
    cab = [c[1] for c in COLUNAS]
    ws.append(cab)
    fill = PatternFill("solid", fgColor="13315C")
    for j, _ in enumerate(cab, 1):
        cell = ws.cell(row=1, column=j)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    def limpa_cel(v):
        return CTRL_RE.sub("", v) if isinstance(v, str) else v

    keys = [c[0] for c in COLUNAS]
    idx_proc = keys.index("link_sei_processo") + 1
    idx_doc = keys.index("link_sei_documento") + 1
    for a in atos:
        linha = [limpa_cel(a.get(c[0], "")) for c in COLUNAS]
        ws.append(linha)
        r = ws.max_row
        # hyperlinks nas colunas de link
        for col_key, col_idx in (("link_sei_processo", idx_proc), ("link_sei_documento", idx_doc)):
            url = a.get(col_key, "")
            if url:
                cell = ws.cell(row=r, column=col_idx)
                cell.hyperlink = url
                cell.value = "abrir SEI"
                cell.font = Font(color="0563C1", underline="single")
    larguras = {"identificador": 26, "ementa": 60, "altera": 28, "revoga": 28,
                "substitui": 28, "cita": 28, "signatario": 26, "tipo": 22,
                "link_sei_processo": 13, "link_sei_documento": 13}
    for j, c in enumerate(COLUNAS, 1):
        ws.column_dimensions[get_column_letter(j)].width = larguras.get(c[0], 12)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    # aba de relações (long format) — ótima para tabela dinâmica
    ws2 = wb.create_sheet("Relações")
    ws2.append(["Ato (origem)", "Relação", "Ato citado", "BS de origem", "Trecho"])
    for j in range(1, 6):
        c = ws2.cell(row=1, column=j)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
    for a in atos:
        for rel in a.get("relacoes", []):
            ws2.append([limpa_cel(a["identificador"]), rel["relacao"],
                        limpa_cel(rel["ato_citado"]), limpa_cel(rel.get("bs_origem", "")),
                        limpa_cel(rel.get("trecho", ""))])
    for j, w in enumerate([28, 18, 30, 18, 70], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    wb.save(caminho)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main():
    ap = argparse.ArgumentParser(description="Extrator do Boletim de Serviço da UFF")
    ap.add_argument("pdfs", nargs="*", help="arquivos PDF (se vazio, usa --pasta)")
    ap.add_argument("--pasta", default=None, help="pasta com PDFs do boletim")
    ap.add_argument("--saida", default=".", help="pasta de saída")
    ap.add_argument("--sem-app", action="store_true",
                    help="não atualiza app/portal-data.json (use em testes/legado)")
    args = ap.parse_args()

    arquivos = list(args.pdfs)
    if not arquivos:
        pasta = args.pasta or os.path.join(os.path.dirname(__file__), "boletins")
        arquivos = sorted(glob.glob(os.path.join(pasta, "*.pdf")))
    if not arquivos:
        print("Nenhum PDF encontrado. Informe arquivos ou --pasta.")
        sys.exit(1)

    todos = []
    metas = []
    falhas = []
    for caminho in arquivos:
        print(f"Lendo {os.path.basename(caminho)} ...", flush=True)
        # Um PDF corrompido/vazio (download truncado pela UFF) não pode
        # derrubar o lote inteiro — pula só esse arquivo e segue os demais.
        try:
            atos, meta = parse_pdf(caminho)
        except Exception as e:
            print(f"   AVISO: pulando {os.path.basename(caminho)} (erro ao ler: {e})")
            falhas.append(os.path.basename(caminho))
            continue
        print(f"   -> {len(atos)} atos | BS {meta['bs_numero']} de {meta['bs_data']}")
        todos.extend(atos)
        metas.append(meta)
    if falhas:
        print(f"\n{len(falhas)} PDF(s) pulado(s) por erro de leitura: {', '.join(falhas)}")

    os.makedirs(args.saida, exist_ok=True)
    salvar_csv(todos, os.path.join(args.saida, "atos.csv"))
    salvar_json(todos, metas, os.path.join(args.saida, "atos.json"))
    salvar_dados_js(todos, metas, os.path.join(args.saida, "dados.js"))
    salvar_xlsx(todos, os.path.join(args.saida, "atos.xlsx"))

    # Atualiza a base do app (Portal de Normas e Atos), se a pasta app/ existir
    # ao lado deste script (ambiente local). No CI isso é feito por gerar_dados_portal.
    app_dir = os.path.join(os.path.dirname(__file__), "app")
    if os.path.isdir(app_dir) and not args.sem_app:
        try:
            import gerar_dados_portal as gdp
            urls = {}
            man = os.path.join(args.pasta or os.path.join(os.path.dirname(__file__), "boletins"),
                               "_urls.json")
            if os.path.exists(man):
                with open(man, encoding="utf-8") as f:
                    urls = json.load(f)
            registros = gdp.converter(_payload(todos, metas), urls)
            destino = os.path.join(app_dir, "portal-data.json")
            with open(destino, "w", encoding="utf-8") as f:
                json.dump(registros, f, ensure_ascii=False)
            print(f"App atualizado: {len(registros)} atos -> app/portal-data.json")
        except Exception as e:
            print(f"  (aviso: não atualizei o app/portal-data.json: {e})")

    n_rel = sum(len(a["relacoes"]) for a in todos)
    n_sei = sum(1 for a in todos if a["processo_sei_principal"])
    print(f"\nTotal: {len(todos)} atos | {n_rel} relações detectadas | "
          f"{n_sei} com processo SEI")
    print(f"Gerados em {os.path.abspath(args.saida)}: atos.csv, atos.json, atos.xlsx")


if __name__ == "__main__":
    main()
